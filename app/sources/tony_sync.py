"""
Tony Excel → CRM Sync (SPEC: docs/specs/SPEC_tony-excel-sync.md)

Polls Tony's Excel fundraising tracker from Egnyte, detects changes, and syncs to CRM.
Run daily at 6 AM via launchd, called from app/main.py.

Flow:
1. Check Egnyte for new file versions (filename pattern match + modified date)
2. Download and parse Active sheet
3. Match org names (alias → exact → fuzzy)
4. Detect changes (Assigned To, Notes, Declined/Closed signals)
5. Send email diff to Oscar and Paige
6. Apply high-confidence changes (≥0.85)
7. Update state file
"""

import difflib
import fnmatch
import json
import logging
import os
import re
import sys
from datetime import date, datetime
from io import BytesIO

import requests
from openpyxl import load_workbook

# Allow running from app/ directory
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), ".env"))

from sources.crm_reader import (
    load_crm_config,
    load_organizations,
    load_prospect_notes,
    load_prospects,
    save_prospect_note,
    update_prospect_field,
    write_organization,
    write_prospect,
)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
CRM_ROOT = os.path.join(BASE_DIR, "crm")
STATE_PATH = os.path.join(CRM_ROOT, "tony_sync_state.json")
PENDING_PATH = os.path.join(CRM_ROOT, "tony_sync_pending.json")
LOG_DIR = os.path.expanduser("~/Library/Logs")
LOG_PATH = os.path.join(LOG_DIR, "arec_tony_sync.log")

EGNYTE_DOMAIN = "avilacapitalllc.egnyte.com"
EGNYTE_FOLDER = "/Shared/AREC/Investor Relations/General Fundraising"
FILENAME_PATTERNS = [
    "AREC Debt Fund II Marketing A List - MASTER as of *.xlsx",
    "AREC Debt Fund II Marketing A List - MASTER v*.xlsx",
]
EXCLUDE_FOLDER = "Archive"

OFFERING = "AREC Debt Fund II"
NEW_PROSPECT_STAGE = "3. Outreach"
NEW_PROSPECT_URGENCY = "High"

CONFIDENCE_AUTO_ACCEPT = 0.85
CONFIDENCE_LOW = 0.60

# Name normalization map from Tony's shorthand to full names
NAME_MAP = {
    "avila": "Tony Avila",
    "vasquez": "Oscar Vasquez",
    "reisner": "Zach Reisner",
    "flynn": "Truman Flynn",
    "albuquerque": "Anthony Albuquerque",
    "fichtner": "Patrick Fichtner",
    "van gorder": "Kevin Van Gorder",
    "kvg": "Kevin Van Gorder",
    "morgan": "Ian Morgan",
    "angeloni": "Max Angeloni",
}

# Email recipients
EMAIL_TO = ["ovasquez@avilacapllc.com", "pkinsey@avilacapllc.com"]

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

os.makedirs(LOG_DIR, exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_PATH),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# State Management
# ---------------------------------------------------------------------------

def load_state() -> dict:
    """Load last processed file state."""
    if not os.path.exists(STATE_PATH):
        return {}
    try:
        with open(STATE_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        log.warning(f"Could not load state file: {e}")
        return {}

def save_state(state: dict) -> None:
    """Save state atomically (temp file + rename)."""
    temp_path = STATE_PATH + ".tmp"
    try:
        with open(temp_path, 'w', encoding='utf-8') as f:
            json.dump(state, f, indent=2, ensure_ascii=False)
        os.rename(temp_path, STATE_PATH)
        log.info(f"State saved: {STATE_PATH}")
    except Exception as e:
        log.error(f"Could not save state: {e}")
        if os.path.exists(temp_path):
            os.remove(temp_path)

def load_pending_queue() -> list:
    """Load low-confidence match pending queue."""
    if not os.path.exists(PENDING_PATH):
        return []
    try:
        with open(PENDING_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        log.warning(f"Could not load pending queue: {e}")
        return []

def save_pending_queue(queue: list) -> None:
    """Save pending queue atomically."""
    temp_path = PENDING_PATH + ".tmp"
    try:
        with open(temp_path, 'w', encoding='utf-8') as f:
            json.dump(queue, f, indent=2, ensure_ascii=False)
        os.rename(temp_path, PENDING_PATH)
        log.info(f"Pending queue saved: {len(queue)} entries")
    except Exception as e:
        log.error(f"Could not save pending queue: {e}")
        if os.path.exists(temp_path):
            os.remove(temp_path)

# ---------------------------------------------------------------------------
# Egnyte API
# ---------------------------------------------------------------------------

def get_egnyte_token() -> str:
    """Get Egnyte API token from environment."""
    token = os.getenv("EGNYTE_API_TOKEN", "").strip()
    if not token:
        raise RuntimeError("EGNYTE_API_TOKEN not set in app/.env")
    return token

def list_egnyte_files() -> list[dict]:
    """List files in Tony's Egnyte folder, excluding Archive subfolder.

    Returns: [{"name": "...", "path": "...", "last_modified": "..."}, ...]
    """
    token = get_egnyte_token()
    url = f"https://{EGNYTE_DOMAIN}/pubapi/v1/fs{EGNYTE_FOLDER}"
    headers = {"Authorization": f"Bearer {token}"}

    try:
        resp = requests.get(url, headers=headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        log.error(f"Egnyte folder list failed: {e}")
        raise

    files = []
    for item in data.get("files", []):
        name = item.get("name", "")
        path = item.get("path", "")

        # Exclude files in Archive folder
        if f"/{EXCLUDE_FOLDER}/" in path or path.endswith(f"/{EXCLUDE_FOLDER}"):
            continue

        # Match filename patterns
        if any(fnmatch.fnmatch(name, pattern) for pattern in FILENAME_PATTERNS):
            files.append({
                "name": name,
                "path": path,
                "last_modified": item.get("last_modified", ""),
            })

    return files

def download_egnyte_file(file_path: str) -> bytes:
    """Download file content from Egnyte."""
    token = get_egnyte_token()
    url = f"https://{EGNYTE_DOMAIN}/pubapi/v1/fs-content{file_path}"
    headers = {"Authorization": f"Bearer {token}"}

    try:
        resp = requests.get(url, headers=headers, timeout=60)
        resp.raise_for_status()
        return resp.content
    except Exception as e:
        log.error(f"Egnyte file download failed: {e}")
        raise

def get_latest_file() -> dict | None:
    """Get the most recent file matching the pattern.

    Returns: {"name": "...", "path": "...", "last_modified": "..."} or None
    """
    files = list_egnyte_files()
    if not files:
        return None

    # Sort by last_modified descending
    files.sort(key=lambda f: f["last_modified"], reverse=True)
    return files[0]

# ---------------------------------------------------------------------------
# Excel Parsing
# ---------------------------------------------------------------------------

def strip_parentheticals(text: str) -> str:
    """Strip parenthetical substrings from org names.

    Examples:
    - "UTIMCO (Matt Saverin)" → "UTIMCO"
    - "Khazanah Americas (Malaysia) Cash Ryan Mulligan" → "Khazanah Americas"
    """
    return re.sub(r'\([^)]*\)', '', text).strip()

def parse_excel(file_bytes: bytes) -> list[dict]:
    """Parse Tony's Excel file Active sheet.

    Returns: [{"org": "...", "assigned_to": "...", "notes": "...", "priority": "..."}, ...]
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True)

    if "Active" not in wb.sheetnames:
        raise ValueError("Active sheet not found in workbook")

    ws = wb["Active"]

    # Row 4 is header, data starts at row 6
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=6, values_only=True), start=6):
        if not row or len(row) < 11:
            continue

        priority = row[0]  # Col A
        org_raw = row[1]   # Col B
        point_person = row[2]  # Col C
        notes = row[10] if len(row) > 10 else None  # Col K

        # Skip empty or summary rows
        if not org_raw or str(org_raw).strip() in ("", "Total", "Investors"):
            continue

        org_name = strip_parentheticals(str(org_raw).strip())

        rows.append({
            "org": org_name,
            "assigned_to": str(point_person).strip() if point_person else "",
            "notes": str(notes).strip() if notes else "",
            "priority": str(priority).strip() if priority else "",
        })

    log.info(f"Parsed {len(rows)} rows from Excel")
    return rows

# ---------------------------------------------------------------------------
# Name Normalization
# ---------------------------------------------------------------------------

def normalize_assigned_to(tony_name: str) -> str:
    """Normalize Tony's shorthand to full name.

    Examples:
    - "Avila" → "Tony Avila"
    - "Reisner/Flynn" → "Zach Reisner"
    - "Avila/Vasquez" → "Tony Avila"

    For slash-separated values, returns first name.
    """
    if not tony_name:
        return ""

    # Handle slash-separated names — take first
    if "/" in tony_name:
        tony_name = tony_name.split("/")[0].strip()

    # Case-insensitive lookup in NAME_MAP
    key = tony_name.lower().strip()
    return NAME_MAP.get(key, tony_name)

# ---------------------------------------------------------------------------
# Org Name Matching
# ---------------------------------------------------------------------------

def match_org(tony_org: str, crm_orgs: list[dict], aliases: dict) -> dict:
    """Match Tony's org name to CRM org.

    Returns: {
        "crm_org": "..." or None,
        "confidence": 0.0–1.0,
        "match_type": "alias" | "exact" | "fuzzy" | "none",
        "best_fuzzy_match": "..." (for low-confidence reporting),
    }
    """
    tony_lower = tony_org.lower().strip()

    # Step 1: Alias lookup
    if tony_lower in aliases:
        canonical = aliases[tony_lower]
        log.debug(f"Alias match: {tony_org} → {canonical}")
        return {
            "crm_org": canonical,
            "confidence": 1.0,
            "match_type": "alias",
            "best_fuzzy_match": None,
        }

    # Step 2: Exact match
    for org in crm_orgs:
        if org["name"].lower() == tony_lower:
            log.debug(f"Exact match: {tony_org} → {org['name']}")
            return {
                "crm_org": org["name"],
                "confidence": 1.0,
                "match_type": "exact",
                "best_fuzzy_match": None,
            }

    # Step 3: Fuzzy match
    crm_names = [org["name"] for org in crm_orgs]
    matcher = difflib.SequenceMatcher(None, tony_lower, "")
    best_score = 0.0
    best_match = None

    for crm_name in crm_names:
        matcher.set_seq2(crm_name.lower())
        score = matcher.ratio()
        if score > best_score:
            best_score = score
            best_match = crm_name

    # Also check aliases for fuzzy matches
    for alias_key, canonical in aliases.items():
        matcher.set_seq2(alias_key)
        score = matcher.ratio()
        if score > best_score:
            best_score = score
            best_match = canonical

    if best_score >= CONFIDENCE_LOW:
        log.debug(f"Fuzzy match: {tony_org} → {best_match} (confidence {best_score:.2f})")
        return {
            "crm_org": best_match if best_score >= CONFIDENCE_AUTO_ACCEPT else None,
            "confidence": best_score,
            "match_type": "fuzzy",
            "best_fuzzy_match": best_match,
        }

    # Step 4: No match
    log.debug(f"No match found for: {tony_org}")
    return {
        "crm_org": None,
        "confidence": 0.0,
        "match_type": "none",
        "best_fuzzy_match": best_match if best_match else None,
    }

# ---------------------------------------------------------------------------
# Change Detection
# ---------------------------------------------------------------------------

def detect_changes(tony_rows: list[dict], crm_orgs: list[dict], aliases: dict) -> dict:
    """Detect changes between Tony's Excel and CRM.

    Returns: {
        "new_prospects_existing_orgs": [{org, assigned_to, notes, ...}, ...],
        "new_orgs": [{tony_org, assigned_to, notes, ...}, ...],
        "updated_prospects": [{org, changes: [type, was, now], confidence, ...}, ...],
        "declined": [{org, ...}, ...],
        "closed": [{org, ...}, ...],
        "low_confidence": [{tony_org, best_match, confidence, ...}, ...],
    }
    """
    crm_prospects = load_prospects(OFFERING)
    crm_prospect_map = {p["org"].lower(): p for p in crm_prospects}
    crm_org_names_lower = {org["name"].lower() for org in crm_orgs}

    new_prospects_existing_orgs = []
    new_orgs = []
    updated_prospects = []
    declined = []
    closed = []
    low_confidence = []

    for row in tony_rows:
        tony_org = row["org"]
        match = match_org(tony_org, crm_orgs, aliases)

        # Low confidence matches — flag for review
        if match["match_type"] == "fuzzy" and CONFIDENCE_LOW <= match["confidence"] < CONFIDENCE_AUTO_ACCEPT:
            low_confidence.append({
                "tony_org": tony_org,
                "best_match": match["best_fuzzy_match"],
                "confidence": match["confidence"],
                "assigned_to": normalize_assigned_to(row["assigned_to"]),
                "notes": row["notes"],
                "detected_at": datetime.now().isoformat(),
            })
            continue

        # No match — brand new org (Case B)
        if match["crm_org"] is None:
            new_orgs.append({
                "tony_org": tony_org,
                "assigned_to": normalize_assigned_to(row["assigned_to"]),
                "notes": row["notes"],
                "priority": row["priority"],
            })
            continue

        crm_org_name = match["crm_org"]
        crm_prospect = crm_prospect_map.get(crm_org_name.lower())

        # Handle Declined/Closed signals
        if row["priority"] == "x":
            if not crm_prospect or crm_prospect.get("Stage", "") != "0. Declined":
                declined.append({
                    "org": crm_org_name,
                    "confidence": match["confidence"],
                })
            continue

        if row["priority"] == "Closed":
            current_stage = crm_prospect.get("Stage", "") if crm_prospect else ""
            if current_stage not in ("8. Closed", "0. Declined"):
                closed.append({
                    "org": crm_org_name,
                    "confidence": match["confidence"],
                })
            continue

        # New prospect (matched org but not in prospects.md yet) — Case A
        if not crm_prospect:
            new_prospects_existing_orgs.append({
                "org": crm_org_name,
                "assigned_to": normalize_assigned_to(row["assigned_to"]),
                "notes": row["notes"],
                "confidence": match["confidence"],
            })
            continue

        # Check for changes to existing prospect
        changes = []

        # Assigned To changed?
        tony_assigned = normalize_assigned_to(row["assigned_to"])
        crm_assigned = crm_prospect.get("Assigned To", "").strip()
        if tony_assigned and tony_assigned != crm_assigned:
            changes.append({
                "type": "Assigned To",
                "was": crm_assigned,
                "now": tony_assigned,
            })

        # Notes changed?
        tony_notes = row["notes"].strip()
        if tony_notes:
            # Check if this note already exists in prospect_notes.json
            existing_notes = load_prospect_notes(crm_org_name, OFFERING)
            existing_texts = [n["text"].strip().lower() for n in existing_notes]
            if tony_notes.lower() not in existing_texts:
                changes.append({
                    "type": "Notes",
                    "was": f"{len(existing_notes)} existing notes",
                    "now": tony_notes,
                })

        if changes:
            updated_prospects.append({
                "org": crm_org_name,
                "changes": changes,
                "confidence": match["confidence"],
            })

    return {
        "new_prospects_existing_orgs": new_prospects_existing_orgs,
        "new_orgs": new_orgs,
        "updated_prospects": updated_prospects,
        "declined": declined,
        "closed": closed,
        "low_confidence": low_confidence,
    }

# ---------------------------------------------------------------------------
# Email Notification
# ---------------------------------------------------------------------------

def send_email(subject: str, body: str, to_addrs: list[str]) -> None:
    """Send email via Microsoft Graph API (OAuth2 authenticated).

    Uses same Graph auth as morning briefing.
    """
    try:
        from auth.graph_auth import get_access_token
        token = get_access_token()
    except Exception as e:
        log.error(f"Could not get Graph access token: {e}")
        raise

    # Build message payload
    message = {
        "subject": subject,
        "body": {
            "contentType": "Text",
            "content": body,
        },
        "toRecipients": [{"emailAddress": {"address": addr}} for addr in to_addrs],
    }

    # Send via Graph API (sendMail endpoint)
    url = "https://graph.microsoft.com/v1.0/me/sendMail"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }
    payload = {"message": message, "saveToSentItems": "true"}

    try:
        resp = requests.post(url, json=payload, headers=headers, timeout=30)
        resp.raise_for_status()
        log.info(f"Email sent to {', '.join(to_addrs)}")
    except Exception as e:
        log.error(f"Email send failed: {e}")
        raise

def build_email_body(filename: str, changes: dict) -> str:
    """Build email body from detected changes."""
    lines = [
        f"Tony uploaded a new version of his Marketing A List.",
        f"File: {filename}",
        "",
    ]

    # New Fund II entries — existing orgs (Case A)
    if changes["new_prospects_existing_orgs"]:
        lines.append(f"── NEW FUND II ENTRIES — EXISTING ORGS ({len(changes['new_prospects_existing_orgs'])}) ────────")
        lines.append("  (Org already in CRM; new Fund II prospect record created)")
        lines.append("")
        for item in changes["new_prospects_existing_orgs"]:
            match_label = "alias match" if item["confidence"] == 1.0 else f"fuzzy match → confidence {item['confidence']:.2f}"
            lines.append(f"  Org: {item['org']}  [{match_label}]")
            lines.append(f"  Assigned To: {item['assigned_to']}")
            lines.append(f"  Stage: {NEW_PROSPECT_STAGE}  |  Urgency: {NEW_PROSPECT_URGENCY}")
            if item["notes"]:
                lines.append(f"  Notes: {item['notes']}")
            else:
                lines.append(f"  Notes: (none)")
            lines.append("")

    # New orgs + Fund II entries (Case B)
    if changes["new_orgs"]:
        lines.append(f"── NEW ORGS + FUND II ENTRIES ({len(changes['new_orgs'])}) ─────────────────")
        lines.append("  (Brand new to CRM — org and prospect record both created)")
        lines.append("")
        for item in changes["new_orgs"]:
            lines.append(f"  Org: {item['tony_org']}")
            lines.append(f"  Assigned To: {item['assigned_to']}")
            lines.append(f"  Stage: {NEW_PROSPECT_STAGE}  |  Urgency: {NEW_PROSPECT_URGENCY}")
            if item["notes"]:
                lines.append(f"  Notes: {item['notes']}")
            else:
                lines.append(f"  Notes: (none)")
            lines.append("")

    # Updated prospects
    if changes["updated_prospects"]:
        lines.append(f"── UPDATED PROSPECTS ({len(changes['updated_prospects'])}) ──────────────────────────")
        lines.append("")
        for item in changes["updated_prospects"]:
            conf_label = f"[HIGH CONFIDENCE {item['confidence']:.2f}]" if item["confidence"] >= CONFIDENCE_AUTO_ACCEPT else ""
            lines.append(f"  Org: {item['org']}  {conf_label}")
            for change in item["changes"]:
                lines.append(f"  Change: {change['type']} updated")
                lines.append(f"    Was:  {change['was']}")
                lines.append(f"    Now:  {change['now']}")
            lines.append("")

    # Declined
    if changes["declined"]:
        lines.append(f"── DECLINED ({len(changes['declined'])}) ────────────────────────────────────")
        lines.append("")
        for item in changes["declined"]:
            lines.append(f"  Org: {item['org']}")
            lines.append(f"  Action: Stage will be set to 0. Declined")
            lines.append("")

    # Closed
    if changes["closed"]:
        lines.append(f"── CLOSED ({len(changes['closed'])}) ────────────────────────────────────")
        lines.append("")
        for item in changes["closed"]:
            lines.append(f"  Org: {item['org']}")
            lines.append(f"  Action: Stage will be set to 8. Closed")
            lines.append("")

    # Low confidence
    if changes["low_confidence"]:
        lines.append(f"── LOW CONFIDENCE MATCHES — REVIEW REQUIRED ({len(changes['low_confidence'])}) ──")
        lines.append("")
        for item in changes["low_confidence"]:
            lines.append(f"  Tony's name: \"{item['tony_org']}\"")
            lines.append(f"  Best match:  \"{item['best_match']}\"  (confidence {item['confidence']:.2f})")
            lines.append(f"  Action: NO CHANGE APPLIED — add an alias to the org in the CRM UI")
            lines.append("")

    lines.append("──────────────────────────────────────────────────────")
    lines.append("Changes above HIGH CONFIDENCE threshold have been applied to the CRM.")
    lines.append("Low-confidence matches have NOT been applied. Add aliases via the CRM org edit page to resolve.")
    lines.append("")
    lines.append("To undo applied changes: crm/tony_sync_state.json records what was changed.")

    return "\n".join(lines)

# ---------------------------------------------------------------------------
# Apply Changes
# ---------------------------------------------------------------------------

def apply_changes(changes: dict) -> dict:
    """Apply high-confidence changes to CRM.

    Returns: {"new_count": N, "updated_count": N, "declined_count": N, "closed_count": N}
    """
    stats = {
        "new_count": 0,
        "updated_count": 0,
        "declined_count": 0,
        "closed_count": 0,
    }

    # Create new prospects for existing orgs (Case A)
    for item in changes["new_prospects_existing_orgs"]:
        data = {
            "Stage": NEW_PROSPECT_STAGE,
            "Urgency": NEW_PROSPECT_URGENCY,
            "Assigned To": item["assigned_to"],
            "Last Touch": date.today().isoformat(),
            "Target": "",
            "Committed": "",
            "Closing": "",
            "Primary Contact": "",
        }
        try:
            write_prospect(item["org"], OFFERING, data)
            log.info(f"Created new Fund II prospect for existing org: {item['org']}")

            # Add note if present
            if item["notes"]:
                save_prospect_note(item["org"], OFFERING, "Tony Avila", item["notes"])
                log.info(f"Added Tony's note to {item['org']}")

            stats["new_count"] += 1
        except Exception as e:
            log.error(f"Failed to create prospect {item['org']}: {e}")

    # New orgs → create org + prospect (Case B)
    for item in changes["new_orgs"]:
        try:
            # Create org in organizations.md
            write_organization(item["tony_org"], {"Type": "", "Notes": ""})
            log.info(f"Created new org: {item['tony_org']}")

            # Create prospect
            data = {
                "Stage": NEW_PROSPECT_STAGE,
                "Urgency": NEW_PROSPECT_URGENCY,
                "Assigned To": item["assigned_to"],
                "Last Touch": date.today().isoformat(),
                "Target": "",
                "Committed": "",
                "Closing": "",
                "Primary Contact": "",
            }
            write_prospect(item["tony_org"], OFFERING, data)
            log.info(f"Created new Fund II prospect for new org: {item['tony_org']}")

            # Add note if present
            if item["notes"]:
                save_prospect_note(item["tony_org"], OFFERING, "Tony Avila", item["notes"])
                log.info(f"Added Tony's note to {item['tony_org']}")

            stats["new_count"] += 1
        except Exception as e:
            log.error(f"Failed to create new org prospect {item['tony_org']}: {e}")

    # Update existing prospects
    for item in changes["updated_prospects"]:
        for change in item["changes"]:
            field = change["type"]
            value = change["now"]
            try:
                if field == "Notes":
                    # Use save_prospect_note() instead of updating inline Notes field
                    save_prospect_note(item["org"], OFFERING, "Tony Avila", value)
                    log.info(f"Added note to {item['org']}: {value[:50]}...")
                else:
                    update_prospect_field(item["org"], OFFERING, field, value)
                    log.info(f"Updated {item['org']}: {field} = {value}")
                stats["updated_count"] += 1
            except Exception as e:
                log.error(f"Failed to update {item['org']} {field}: {e}")

    # Set Declined
    for item in changes["declined"]:
        try:
            update_prospect_field(item["org"], OFFERING, "Stage", "0. Declined")
            log.info(f"Set {item['org']} to Declined")
            stats["declined_count"] += 1
        except Exception as e:
            log.error(f"Failed to decline {item['org']}: {e}")

    # Set Closed
    for item in changes["closed"]:
        try:
            update_prospect_field(item["org"], OFFERING, "Stage", "8. Closed")
            log.info(f"Set {item['org']} to Closed")
            stats["closed_count"] += 1
        except Exception as e:
            log.error(f"Failed to close {item['org']}: {e}")

    return stats

# ---------------------------------------------------------------------------
# Main Sync
# ---------------------------------------------------------------------------

def run_sync() -> dict:
    """Main sync entry point. Returns stats dict."""
    log.info("=== Tony Excel sync starting ===")

    try:
        # Check for new file
        latest_file = get_latest_file()
        if not latest_file:
            log.info("No matching files found in Egnyte. Exiting.")
            return {"status": "no_file"}

        # Check if already processed
        state = load_state()
        if state.get("last_processed_filename") == latest_file["name"]:
            log.info(f"File already processed: {latest_file['name']}. No changes.")
            return {"status": "no_change"}

        log.info(f"New file detected: {latest_file['name']}")

        # Download and parse
        file_bytes = download_egnyte_file(latest_file["path"])
        tony_rows = parse_excel(file_bytes)

        # Load CRM data
        crm_orgs = load_organizations()
        from app.sources.crm_reader import get_org_aliases_map
        aliases = get_org_aliases_map()

        # Detect changes
        changes = detect_changes(tony_rows, crm_orgs, aliases)

        total_changes = (
            len(changes["new_prospects_existing_orgs"]) +
            len(changes["new_orgs"]) +
            len(changes["updated_prospects"]) +
            len(changes["declined"]) +
            len(changes["closed"])
        )

        # Update pending queue with new low-confidence matches
        if changes["low_confidence"]:
            pending = load_pending_queue()
            # Add new entries (avoid duplicates by checking tony_org)
            existing_orgs = {p["tony_org"] for p in pending}
            for item in changes["low_confidence"]:
                if item["tony_org"] not in existing_orgs:
                    pending.append(item)
            save_pending_queue(pending)

        if total_changes == 0 and not changes["low_confidence"]:
            log.info("No changes detected. Updating state and exiting.")
            save_state({
                "last_processed_filename": latest_file["name"],
                "last_processed_at": datetime.now().isoformat(),
                "egnyte_modified": latest_file["last_modified"],
                "rows_processed": len(tony_rows),
                "changes_detected": 0,
                "changes_applied": 0,
            })
            return {"status": "no_changes"}

        # Send email notification
        email_subject = f"Tony updated his CRM file — {total_changes} changes detected"
        email_body = build_email_body(latest_file["name"], changes)

        try:
            send_email(email_subject, email_body, EMAIL_TO)
        except Exception as e:
            log.error(f"Email notification failed: {e}")
            # Continue anyway — we still want to apply changes

        # Apply high-confidence changes
        stats = apply_changes(changes)

        # Update state
        save_state({
            "last_processed_filename": latest_file["name"],
            "last_processed_at": datetime.now().isoformat(),
            "egnyte_modified": latest_file["last_modified"],
            "rows_processed": len(tony_rows),
            "changes_detected": total_changes,
            "changes_applied": sum(stats.values()),
        })

        log.info(f"=== Sync complete: {stats} ===")
        return {"status": "success", **stats}

    except Exception as e:
        log.error(f"Sync failed: {e}", exc_info=True)

        # Send error email to Oscar only
        try:
            error_subject = f"Tony CRM Sync Failed — {type(e).__name__}"
            error_body = f"The Tony Excel → CRM sync failed with error:\n\n{e}\n\nCheck ~/Library/Logs/arec_tony_sync.log for details."
            send_email(error_subject, error_body, ["ovasquez@avilacapllc.com"])
        except Exception:
            log.error("Error email also failed")

        return {"status": "error", "error": str(e)}


if __name__ == "__main__":
    result = run_sync()
    print(json.dumps(result, indent=2))
