"""
crm_blueprint.py — Flask CRM blueprint (/crm routes).
Extracted from dashboard.py for maintainability.
"""

import os
import sys
import re
import json
import hashlib
import anthropic
from datetime import date, datetime
from urllib.parse import quote as urlquote
from flask import (
    Blueprint, jsonify, request, render_template,
    redirect, url_for, abort, send_file, g,
)

_APP_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _APP_DIR not in sys.path:
    sys.path.insert(0, _APP_DIR)

PROJECT_ROOT = os.path.dirname(_APP_DIR)
TASKS_PATH = os.path.join(PROJECT_ROOT, "TASKS.md")

# Simple pass-through decorator for local dev (no auth needed)
def login_required(f):
    """No-op decorator for local dev."""
    return f

from sources.crm_reader import (
    load_prospects, load_offerings, get_fund_summary, get_fund_summary_all,
    load_crm_config, get_organization, write_organization, load_organizations,
    get_contacts_for_org, create_person_file, update_contact_fields,
    get_prospects_for_org, get_prospect, write_prospect, update_prospect_field,
    load_unmatched, remove_unmatched, add_unmatched,
    _parse_currency, load_person, load_all_persons, load_tasks_by_org,
    delete_prospect, load_meeting_history, add_meeting_entry,
    get_tasks_for_prospect as get_tasks_for_prospect_db,
    get_all_prospect_tasks as get_all_tasks_for_dashboard,
    get_all_prospect_tasks_with_index,
    add_prospect_task, complete_prospect_task,
    get_tasks_grouped_by_prospect, get_tasks_grouped_by_owner,
    load_email_log, get_emails_for_org, find_email_by_message_id,
    load_interactions, append_interaction,
    save_brief, load_saved_brief, load_all_briefs,
    load_prospect_notes, save_prospect_note,
    load_org_notes, save_org_note,
    load_prospect_meetings, save_prospect_meeting, delete_prospect_meeting,
    load_meetings, get_meeting, save_meeting, update_meeting, delete_meeting,
    process_meeting_notes, approve_meeting_insight, dismiss_meeting_insight,
    append_person_email_history, append_org_email_history,
    discover_and_enrich_contact_emails, enrich_org_domain,
    find_person_by_email,
    get_merge_preview, merge_organizations,
    get_primary_contact, set_primary_contact, clear_primary_contact,
    resolve_org_name,
    normalize_team_name,
    get_heatmap_prospects,
)
from sources.memory_reader import _parse_task_line, _format_task_line
from sources.relationship_brief import (
    find_people_files, find_glossary_entry, find_meeting_summaries, find_org_tasks,
    collect_relationship_data, build_context_block, build_fallback_summary,
    compute_content_hash, BRIEF_SYSTEM_PROMPT, PROSPECT_BRIEF_SYSTEM_PROMPT,
    collect_person_data, build_person_context_block, build_person_fallback_summary,
    execute_person_updates, PERSON_BRIEF_SYSTEM_PROMPT, PERSON_UPDATE_ROUTING_PROMPT,
)
from briefing.brief_synthesizer import call_claude_brief

crm_bp = Blueprint('crm', __name__, url_prefix='/crm')


@crm_bp.context_processor
def inject_search_index():
    """Inject global search index into every CRM template."""
    try:
        entries = []
        for p in load_prospects():
            entries.append({
                'name': p['org'],
                'secondary': p['offering'],
                'type': 'prospect',
                'typeLabel': 'Prospect',
                'url': f"/crm/prospect/{urlquote(p['offering'], safe='')}/{urlquote(p['org'], safe='')}/detail",
            })
        for person in load_all_persons():
            entries.append({
                'name': person['name'],
                'secondary': person.get('organization', ''),
                'type': 'person',
                'typeLabel': 'Person',
                'url': f"/crm/people/{person['slug']}",
            })
        for org in load_organizations():
            entries.append({
                'name': org['name'],
                'secondary': org.get('Aliases', ''),
                'type': 'org',
                'typeLabel': 'Org',
                'url': f"/crm/org/{urlquote(org['name'], safe='')}",
            })
        return {'search_index_json': json.dumps(entries)}
    except Exception:
        return {'search_index_json': '[]'}


EDITABLE_FIELDS = {
    'stage', 'urgent', 'target', 'assigned_to', 'notes', 'closing'
}

ORG_BRIEF_SYSTEM = (
    "You are a senior relationship intelligence analyst for Avila Real Estate Capital (AREC), "
    "a private real estate credit fund manager. Write a concise 2-3 paragraph organizational "
    "brief summarizing: who this organization is, their relationship with AREC, current status, "
    "key contacts, and any open opportunities or risks. Be specific and actionable. "
    "Write in plain prose — no headers, no bullets."
)


# ---------------------------------------------------------------------------
# KB people helper
# ---------------------------------------------------------------------------

def parse_kb_person_file(path: str) -> dict:
    """Parse a memory/people/*.md file into a person dict."""
    slug = os.path.splitext(os.path.basename(path))[0]
    name_fallback = ' '.join(w.capitalize() for w in slug.replace('-', ' ').split())
    fields = {'name': name_fallback, 'org': '', 'title': '', 'email': '', 'phone': '', 'company': ''}
    field_map = {
        'name':         'name',
        'organization': 'company',
        'org':          'company',
        'company':      'company',
        'title':        'title',
        'role':         'title',
        'email':        'email',
        'phone':        'phone',
        'cell':         'phone',
        'mobile':       'phone',
    }
    try:
        with open(path, 'r', encoding='utf-8') as f:
            for line in f:
                h1 = re.match(r'^#\s+(.+)', line.strip())
                if h1:
                    fields['name'] = h1.group(1).strip()
                    continue
                m = re.match(r'(?:-|\*)?\s*\*\*([^:]+):\*\*\s*(.*)', line.strip())
                if m:
                    key = m.group(1).lower().strip()
                    val = m.group(2).strip()
                    if key in field_map and val:
                        fields[field_map[key]] = val
    except Exception:
        pass
    fields['slug'] = slug
    fields['org'] = fields['company']
    return fields


# ---------------------------------------------------------------------------
# KB people API
# ---------------------------------------------------------------------------

@crm_bp.route('/api/kb-people')
@login_required
def api_kb_people():
    q = request.args.get('q', '').lower().strip()
    config = load_crm_config()
    team = config.get('team', [])
    arec_team = {(m['name'] if isinstance(m, dict) else m).lower() for m in team}
    contacts_dir = os.path.join(PROJECT_ROOT, 'contacts')
    kb_dir = os.path.join(PROJECT_ROOT, 'memory', 'people')
    results = []
    seen_slugs = set()
    for people_dir in [contacts_dir, kb_dir]:
        if not os.path.isdir(people_dir):
            continue
        for fname in sorted(os.listdir(people_dir)):
            if not fname.endswith('.md'):
                continue
            slug = fname[:-3]
            if slug in seen_slugs:
                continue
            seen_slugs.add(slug)
            path = os.path.join(people_dir, fname)
            person = parse_kb_person_file(path)
            if person['name'].lower() in arec_team:
                continue
            if not q or q in person['name'].lower():
                results.append(person)
    results.sort(key=lambda p: p['name'].lower())
    return jsonify(results[:10] if q else results)


@crm_bp.route('/api/people/search')
@login_required
def api_people_search():
    """Alias for /api/kb-people for spec compliance."""
    return api_kb_people()


# ---------------------------------------------------------------------------
# Page routes
# ---------------------------------------------------------------------------

@crm_bp.route('/person/<slug>')
@login_required
def person_detail(slug):
    contacts_dir = os.path.join(PROJECT_ROOT, 'contacts')
    kb_dir = os.path.join(PROJECT_ROOT, 'memory', 'people')
    path = os.path.join(contacts_dir, f'{slug}.md')
    if not os.path.exists(path):
        path = os.path.join(kb_dir, f'{slug}.md')
    if not os.path.exists(path):
        abort(404)
    person = parse_kb_person_file(path)
    with open(path, 'r', encoding='utf-8') as f:
        raw_content = f.read()
    emails = get_emails_for_org(person['org']) if person.get('org') else []
    if person.get('email'):
        em = person['email'].lower()
        person_emails = [
            e for e in emails
            if em in e.get('from', '').lower()
            or em in ' '.join(e.get('to', [])).lower()
            or em in ' '.join(e.get('cc', [])).lower()
        ]
    else:
        person_emails = []
    return render_template('crm_person_detail.html',
                           person=person,
                           slug=slug,
                           raw_content=raw_content,
                           person_emails=person_emails[:20])


@crm_bp.route('/')
@crm_bp.route('')
@login_required
def pipeline():
    config = load_crm_config()
    offerings = load_offerings()
    return render_template('crm_pipeline.html', config=config, offerings=offerings)


@crm_bp.route('/people')
@login_required
def people_list():
    return render_template('crm_people.html')


@crm_bp.route('/people/<slug>')
@login_required
def people_person_detail(slug):
    return _render_person_detail(slug)


def _render_person_detail(slug):
    contacts_dir = os.path.join(PROJECT_ROOT, 'contacts')
    kb_dir = os.path.join(PROJECT_ROOT, 'memory', 'people')
    path = os.path.join(contacts_dir, f'{slug}.md')
    if not os.path.exists(path):
        path = os.path.join(kb_dir, f'{slug}.md')
    if not os.path.exists(path):
        abort(404)
    person = parse_kb_person_file(path)
    return render_template('crm_person_detail.html', person=person, slug=slug)


@crm_bp.route('/people/<slug>/delete', methods=['POST'])
@login_required
def delete_person(slug):
    contacts_dir = os.path.join(PROJECT_ROOT, 'contacts')
    kb_dir = os.path.join(PROJECT_ROOT, 'memory', 'people')
    path = os.path.join(contacts_dir, f'{slug}.md')
    if not os.path.exists(path):
        path = os.path.join(kb_dir, f'{slug}.md')
    if not os.path.exists(path):
        abort(404)
    os.remove(path)

    # Remove from contacts_index.md
    index_path = os.path.join(PROJECT_ROOT, 'crm', 'contacts_index.md')
    if os.path.exists(index_path):
        with open(index_path, 'r') as f:
            lines = f.readlines()
        filtered = [l for l in lines if slug not in l]
        with open(index_path, 'w') as f:
            f.writelines(filtered)

    return ('', 204)


@crm_bp.route('/orgs')
@login_required
def orgs_list():
    config = load_crm_config()
    return render_template('crm_orgs.html', config=config)


@crm_bp.route('/org/<path:name>/edit')
@login_required
def org_edit(name):
    config = load_crm_config()
    offerings = load_offerings()
    org_data = get_organization(name) or {'name': name, 'Type': '', 'Notes': ''}
    contacts = get_contacts_for_org(name)
    prospects = get_prospects_for_org(name)

    # Load org notes
    org_notes = load_org_notes(name)

    # Load meetings (org-owned)
    meetings = load_meeting_history(name)

    # Load emails (org-owned)
    emails = get_emails_for_org(name)

    org_brief_saved = load_saved_brief('org', name)

    return render_template('crm_org_edit.html',
                           org_name=name,
                           config=config,
                           offerings=offerings,
                           org_data=org_data,
                           contacts=contacts,
                           prospects=prospects,
                           org_notes=org_notes,
                           meetings=meetings,
                           emails=emails,
                           org_brief_saved=org_brief_saved)


@crm_bp.route('/org/<path:name>')
@login_required
def org_detail(name):
    """Redirect to org edit page for backward compatibility."""
    return redirect(url_for('crm.org_edit', name=name))


@crm_bp.route('/prospect/<offering>/<path:org>')
@login_required
def prospect_edit(offering, org):
    prospect = get_prospect(org, offering)
    if not prospect:
        abort(404)
    config = load_crm_config()
    contacts = get_contacts_for_org(org)
    arec_names = {member['name'].lower() for member in config['team']}
    org_slugs = {c['slug'] for c in contacts}
    other_contacts = [
        p for p in load_all_persons()
        if p['slug'] not in org_slugs
        and p['name'].lower() not in arec_names
        and 'arec' not in (p.get('organization') or '').lower()
        and 'avila real estate' not in (p.get('organization') or '').lower()
    ]
    return render_template('crm_prospect_edit.html',
                           prospect=prospect,
                           config=config,
                           contacts=contacts,
                           other_contacts=other_contacts,
                           offering=offering,
                           org=org)


@crm_bp.route('/prospect/<offering>/<path:org>/detail')
@login_required
def prospect_detail(offering, org):
    prospect = get_prospect(org, offering)
    if not prospect:
        abort(404)
    config = load_crm_config()
    urgent_raw = prospect.get('Urgent', '') or prospect.get('urgent', '')
    prospect['urgent_bool'] = str(urgent_raw).strip().lower() in ('yes', 'true', 'high', '1')

    # Load org data for the org sub-section
    org_data = get_organization(org) or {'name': org, 'Type': ''}
    contacts = get_contacts_for_org(org)
    primary_contact_name = prospect.get('Primary Contact', '')

    # Load org brief for read-only display
    org_brief_saved = load_saved_brief('org', org)

    # Load prospect brief for editable display
    prospect_brief_key = f"{org}::{offering}"
    prospect_brief_saved = load_saved_brief('prospect', prospect_brief_key)

    # Load meetings (org-owned, displayed on both pages)
    meetings = load_meeting_history(org)

    # Load emails (org-owned, displayed on both pages)
    emails = get_emails_for_org(org)

    return render_template('crm_prospect_detail.html',
                           prospect=prospect,
                           config=config,
                           offering=offering,
                           org=org,
                           org_data=org_data,
                           contacts=contacts,
                           primary_contact_name=primary_contact_name,
                           org_brief_saved=org_brief_saved,
                           prospect_brief_saved=prospect_brief_saved,
                           meetings=meetings,
                           emails=emails)


# ---------------------------------------------------------------------------
# Brief API
# ---------------------------------------------------------------------------

def _run_prospect_brief(org: str, offering: str, max_tokens: int = 1600,
                        want_json: bool = True) -> tuple:
    """Collect context, call Claude, and persist a prospect brief.

    Returns (narrative, at_a_glance, content_hash).
    """
    try:
        raw_data = collect_relationship_data(org, offering, base_dir=PROJECT_ROOT)
    except Exception as e:
        print(f"[brief] collect_relationship_data failed for {org}: {e}")
        raw_data = {}
    content_hash = compute_content_hash(raw_data)
    context_block = build_context_block(raw_data)
    narrative, at_a_glance = call_claude_brief(
        BRIEF_SYSTEM_PROMPT,
        f"Generate a relationship brief for {org} regarding {offering}.\n\n{context_block}",
        max_tokens=max_tokens,
        want_json=want_json,
    )
    brief_key = f"{org}::{offering}"
    save_brief('prospect', brief_key, narrative, content_hash, at_a_glance=at_a_glance)
    return narrative, at_a_glance, content_hash


@crm_bp.route('/api/prospect/<offering>/<path:org>/brief', methods=['GET', 'POST'])
@login_required
def api_prospect_brief(offering, org):
    if request.method == 'GET':
        try:
            raw_data = collect_relationship_data(org, offering, base_dir=PROJECT_ROOT)
        except Exception as e:
            print(f"[brief] GET collect_relationship_data failed for {org}: {e}")
            raw_data = {}
        try:
            content_hash = compute_content_hash(raw_data)
            brief_key = f"{org}::{offering}"
            saved = load_saved_brief('prospect', brief_key)
            prospect = raw_data.get('prospect', {})
            return jsonify({
                **raw_data,
                'content_hash': content_hash,
                'saved_brief': saved,
                'relationship_brief': prospect.get('Relationship Brief', ''),
                'brief_refreshed': prospect.get('Brief Refreshed', ''),
            })
        except Exception as e:
            print(f"[brief] GET response build failed for {org}: {e}")
            import traceback; traceback.print_exc()
            return jsonify({'error': str(e)}), 500

    # POST — synthesize and persist
    try:
        today_str = date.today().isoformat()
        narrative, at_a_glance, _ = _run_prospect_brief(org, offering)
        brief_text = narrative.replace('\n', ' ').strip()
        update_prospect_field(org, offering, 'relationship_brief', brief_text)
        update_prospect_field(org, offering, 'brief_refreshed', today_str)
        return jsonify({
            'narrative': narrative,
            'brief_refreshed': today_str,
            'at_a_glance': at_a_glance,
        })
    except Exception as e:
        print(f"[brief] POST synthesis failed for {org}/{offering}: {e}")
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e), 'narrative': '', 'at_a_glance': []}), 500


@crm_bp.route('/api/synthesize-brief', methods=['POST'])
@login_required
def api_synthesize_brief():
    """Call Claude API to synthesize a narrative relationship brief from raw data."""
    data = request.get_json(force=True)
    org = data.get('org', '').strip()
    offering = data.get('offering', '').strip()
    generate_glance = bool(data.get('generate_glance', False))
    if not org or not offering:
        return jsonify({'error': 'org and offering required'}), 400
    narrative, at_a_glance, content_hash = _run_prospect_brief(
        org, offering,
        max_tokens=1600 if generate_glance else 1500,
        want_json=generate_glance,
    )
    return jsonify({
        'narrative': narrative,
        'content_hash': content_hash,
        'at_a_glance': at_a_glance,
    })


def _run_focused_prospect_brief(org: str, offering: str) -> tuple:
    """Generate a short, offering-specific prospect brief.
    Returns (narrative, at_a_glance, content_hash).
    """
    try:
        raw_data = collect_relationship_data(org, offering, base_dir=PROJECT_ROOT)
    except Exception as e:
        print(f"[prospect-brief] collect_relationship_data failed for {org}: {e}")
        raw_data = {}
    content_hash = compute_content_hash(raw_data)
    context_block = build_context_block(raw_data)
    narrative, at_a_glance = call_claude_brief(
        PROSPECT_BRIEF_SYSTEM_PROMPT,
        f"Generate a concise prospect brief for {offering} at {org}.\n\n{context_block}",
        max_tokens=800,
        want_json=True,
    )
    brief_key = f"{org}::{offering}"
    save_brief('prospect', brief_key, narrative, content_hash, at_a_glance=at_a_glance)
    return narrative, at_a_glance, content_hash


@crm_bp.route('/api/prospect/<offering>/<path:org>/prospect-brief', methods=['GET', 'POST'])
@login_required
def api_prospect_brief_focused(offering, org):
    """Prospect-specific brief (short, offering-focused status)."""
    if request.method == 'GET':
        brief_key = f"{org}::{offering}"
        saved = load_saved_brief('prospect', brief_key)
        return jsonify(saved or {})

    # POST — synthesize and persist
    try:
        narrative, at_a_glance, content_hash = _run_focused_prospect_brief(org, offering)
        return jsonify({
            'narrative': narrative,
            'at_a_glance': at_a_glance,
            'content_hash': content_hash,
            'generated_at': datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ'),
        })
    except Exception as e:
        print(f"[prospect-brief] POST synthesis failed for {org}/{offering}: {e}")
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e), 'narrative': ''}), 500


# ---------------------------------------------------------------------------
# Email API
# ---------------------------------------------------------------------------

@crm_bp.route('/api/emails/<path:org>')
@login_required
def api_emails_for_org(org):
    """Return paginated emails for an org from email_log.json."""
    limit = request.args.get('limit', 20, type=int)
    offset = request.args.get('offset', 0, type=int)
    emails = get_emails_for_org(org)
    paginated = emails[offset:offset + limit]
    return jsonify({
        'emails': paginated,
        'total': len(emails),
        'offset': offset,
        'limit': limit,
    })


@crm_bp.route('/api/email/<path:message_id>')
@login_required
def api_email_detail(message_id):
    """Return a single email entry from the log."""
    email = find_email_by_message_id(message_id)
    if not email:
        abort(404)
    return jsonify(email)


# ---------------------------------------------------------------------------
# Prospect Notes
# ---------------------------------------------------------------------------

@crm_bp.route('/api/prospect/<offering>/<path:org>/add-note', methods=['POST'])
@login_required
def api_add_prospect_note(offering, org):
    data = request.get_json(force=True)
    # Author is auto-populated from the logged-in user
    author = (g.user.get('display_name') or g.user.get('email') or 'Unknown').strip()
    text = (data.get('text') or '').strip()
    if not text:
        return jsonify({'error': 'text is required'}), 400
    entry = save_prospect_note(org, offering, author, text)
    # Return full updated notes list so the client can render immediately
    all_notes = load_prospect_notes(org, offering)
    return jsonify({'ok': True, 'entry': entry, 'notes_log': all_notes})


# ---------------------------------------------------------------------------
# Upcoming Meetings API
# ---------------------------------------------------------------------------

# Prospect upcoming meetings API removed — feature moved to calendar integration
# @crm_bp.route('/api/prospect/<offering>/<path:org>/add-meeting', methods=['POST'])
# @crm_bp.route('/api/prospect/<offering>/<path:org>/delete-meeting', methods=['POST'])


# ---------------------------------------------------------------------------
# Person API
# ---------------------------------------------------------------------------

@crm_bp.route('/api/person-data')
@login_required
def api_person_data():
    name = request.args.get('name', '').strip()
    if not name:
        return jsonify({'error': 'name required'}), 400
    data = collect_person_data(name, base_dir=PROJECT_ROOT)
    content_hash = hashlib.md5(
        json.dumps(data, sort_keys=True, default=str).encode()
    ).hexdigest()[:12]
    saved = load_saved_brief('person', name)

    # Augment profile with DB enrichment fields (linkedin_url, enriched_at)
    slug = re.sub(r'[^a-z0-9\s-]', '', name.lower().strip())
    slug = re.sub(r'\s+', '-', slug).strip('-')
    db_person = load_person(slug)
    if db_person:
        data['profile']['linkedin_url'] = db_person.get('linkedin_url') or ''
        data['profile']['enriched_at'] = db_person.get('enriched_at')

    return jsonify({**data, 'content_hash': content_hash, 'saved_brief': saved})


@crm_bp.route('/api/synthesize-person-brief', methods=['POST'])
@login_required
def api_synthesize_person_brief():
    """Synthesize a person-focused AI narrative brief."""
    data = request.get_json(force=True)
    person_name = data.get('name', '').strip()
    if not person_name:
        return jsonify({'error': 'name required'}), 400
    raw_data = collect_person_data(person_name, base_dir=PROJECT_ROOT)
    content_hash = hashlib.md5(
        json.dumps(raw_data, sort_keys=True, default=str).encode()
    ).hexdigest()[:12]
    context_block = build_person_context_block(raw_data)
    try:
        narrative, _ = call_claude_brief(
            PERSON_BRIEF_SYSTEM_PROMPT,
            f"Generate a person brief for {person_name}.\n\n{context_block}",
            max_tokens=1500,
            want_json=False,
        )
    except Exception:
        narrative = build_person_fallback_summary(raw_data)
    save_brief('person', person_name, narrative, content_hash)
    return jsonify({
        'narrative': narrative,
        'content_hash': content_hash,
    })


@crm_bp.route('/api/person-update', methods=['POST'])
@login_required
def api_person_update():
    """Accept free-text context about a person, AI routes updates to data stores."""
    data = request.get_json(force=True)
    person_name = data.get('name', '').strip()
    user_input = data.get('input', '').strip()
    if not person_name or not user_input:
        return jsonify({'error': 'name and input required'}), 400
    raw_data = collect_person_data(person_name, base_dir=PROJECT_ROOT)
    context_block = build_person_context_block(raw_data)
    config = load_crm_config()
    org_name = raw_data.get('org_name', '')
    try:
        client = anthropic.Anthropic()
        message = client.messages.create(
            model='claude-sonnet-4-6',
            max_tokens=2000,
            system=PERSON_UPDATE_ROUTING_PROMPT,
            messages=[{
                'role': 'user',
                'content': (
                    f"CURRENT PERSON DATA:\n{context_block}\n\n"
                    f"PERSON NAME: {person_name}\n"
                    f"ORGANIZATION: {org_name}\n"
                    f"VALID STAGES: {', '.join(config.get('stages', []))}\n"
                    f"TODAY'S DATE: {date.today().isoformat()}\n\n"
                    f"USER UPDATE:\n{user_input}\n\n"
                    "Determine what data store updates are needed and return JSON."
                )
            }]
        )
        response_text = message.content[0].text
        clean = response_text.replace('```json', '').replace('```', '').strip()
        updates = json.loads(clean)
    except Exception as e:
        return jsonify({'error': f'AI routing failed: {str(e)}'}), 500
    results = execute_person_updates(person_name, org_name, updates, base_dir=PROJECT_ROOT)
    updated_data = collect_person_data(person_name, base_dir=PROJECT_ROOT)
    new_hash = hashlib.md5(
        json.dumps(updated_data, sort_keys=True, default=str).encode()
    ).hexdigest()[:12]
    return jsonify({'actions': results, 'new_content_hash': new_hash})


@crm_bp.route('/people/api/<slug>/contact', methods=['PATCH'])
@login_required
def api_person_contact_update(slug):
    data = request.get_json(force=True)
    contacts_dir = os.path.join(PROJECT_ROOT, 'contacts')
    kb_dir = os.path.join(PROJECT_ROOT, 'memory', 'people')
    path = os.path.join(contacts_dir, f'{slug}.md')
    if not os.path.exists(path):
        path = os.path.join(kb_dir, f'{slug}.md')
    if not os.path.exists(path):
        return jsonify({'error': 'Person not found'}), 404

    FIELD_ORDER = ['Company', 'Title', 'Email', 'Phone']
    FIELD_MAP = {
        'company': 'Company', 'organization': 'Company', 'org': 'Company',
        'title': 'Title', 'role': 'Title',
        'email': 'Email',
        'phone': 'Phone', 'cell': 'Phone', 'mobile': 'Phone',
    }
    CONTACT_RE = re.compile(r'^-?\s*\*\*([^:]+):\*\*\s*(.*)', re.IGNORECASE)

    new_values = {
        'Company': data.get('company', '').strip(),
        'Title':   data.get('title', '').strip(),
        'Email':   data.get('email', '').strip(),
        'Phone':   data.get('phone', '').strip(),
    }

    with open(path, 'r', encoding='utf-8') as f:
        raw = [ln.rstrip('\n') for ln in f.readlines()]

    # 1. Detect optional h1 heading at top of file
    h1_line = None
    scan_start = 0
    if raw and re.match(r'^#\s', raw[0]):
        h1_line = raw[0]
        scan_start = 1

    # 2. Find extent of existing contact block (lines before first ## heading or prose)
    existing_fields = {}
    block_end = scan_start
    for i in range(scan_start, len(raw)):
        stripped = raw[i].strip()
        if not stripped:
            continue
        if stripped.startswith('##'):
            break
        m = CONTACT_RE.match(stripped)
        if m and m.group(1).lower().strip() in FIELD_MAP:
            canonical = FIELD_MAP[m.group(1).lower().strip()]
            existing_fields.setdefault(canonical, m.group(2).strip())
            block_end = i + 1
            continue
        break  # non-blank, non-field line = end of block

    # 3. Merge new values into existing, then build the new contact block
    existing_fields.update(new_values)
    new_block = [
        f'- **{field}:** {existing_fields[field]}'
        for field in FIELD_ORDER
        if existing_fields.get(field)
    ]

    # 4. Everything after the old block, stripping leading blank lines
    rest = raw[block_end:]
    while rest and not rest[0].strip():
        rest.pop(0)

    # 5. Assemble new file content
    parts = []
    if h1_line is not None:
        parts.append(h1_line)
    if new_block:
        if parts:
            parts.append('')  # blank line after h1
        parts.extend(new_block)
    if rest:
        if parts:
            parts.append('')  # blank line before rest
        parts.extend(rest)

    new_content = '\n'.join(parts)
    if new_content and not new_content.endswith('\n'):
        new_content += '\n'

    with open(path, 'w', encoding='utf-8') as f:
        f.write(new_content)

    person = parse_kb_person_file(path)
    return jsonify(person)


# ---------------------------------------------------------------------------
# Contact Enrichment
# ---------------------------------------------------------------------------

def _search_linkedin_url(name: str, org: str, timeout: int = 10) -> str | None:
    """
    Search DuckDuckGo for a LinkedIn profile URL for the given person + org.
    Returns the first matching linkedin.com/in/ URL, or None.
    """
    import requests as _requests
    from urllib.parse import unquote, urlencode

    query = f'"{name}" "{org}" site:linkedin.com/in'
    try:
        resp = _requests.get(
            'https://html.duckduckgo.com/html/',
            params={'q': query},
            headers={'User-Agent': 'Mozilla/5.0 (compatible; AREC-CRM/1.0)'},
            timeout=timeout,
        )
        if resp.status_code != 200:
            return None
        # DuckDuckGo HTML encodes result URLs in uddg= parameters
        matches = re.findall(r'uddg=(https?%3A%2F%2F(?:www\.)?linkedin\.com%2Fin%2F[\w\-]+)', resp.text)
        if matches:
            return unquote(matches[0])
    except Exception as e:
        print(f"[enrich] LinkedIn search failed: {e}")
    return None


@crm_bp.route('/api/people/<slug>/enrich', methods=['POST'])
@login_required
def api_person_enrich(slug):
    """
    Run the contact enrichment pipeline for a person:
      1. LinkedIn URL via web search
      2. Email footer parsing (phone, title) from recent emails
      3. Outlook contacts lookup (phone, title, email)

    Returns a JSON preview of discovered fields — does NOT auto-save.
    """
    person = load_person(slug)
    if not person:
        return jsonify({'error': 'Person not found'}), 404

    current = {
        'phone': person.get('phone') or '',
        'title': person.get('role') or '',
        'email': person.get('email') or '',
        'linkedin_url': person.get('linkedin_url') or '',
    }

    findings = []
    graph_error = None

    # --- 1. LinkedIn search ---
    if person.get('organization'):
        linkedin_url = _search_linkedin_url(person['name'], person['organization'])
        if linkedin_url:
            findings.append({
                'field': 'linkedin_url',
                'label': 'LinkedIn',
                'found': linkedin_url,
                'current': current['linkedin_url'],
                'status': 'CONFIRMED' if linkedin_url == current['linkedin_url']
                          else ('NEW' if not current['linkedin_url'] else 'CONFLICT'),
                'source': 'web search',
            })

    # --- 2 & 3. Graph API: email footer + Outlook contacts ---
    try:
        from auth.graph_auth import get_access_token
        from sources.ms_graph import search_contact_emails_for_signature, lookup_outlook_contact

        token = get_access_token(allow_device_flow=False)

        # Email footer scan
        if person.get('email'):
            sig = search_contact_emails_for_signature(
                token, person['name'], person['email'], days_back=30
            )
            if sig.get('phone'):
                source_label = f"email footer ({sig['email_count']} email{'s' if sig['email_count'] != 1 else ''})"
                findings.append({
                    'field': 'phone',
                    'label': 'Phone',
                    'found': sig['phone'],
                    'current': current['phone'],
                    'status': 'CONFIRMED' if sig['phone'] == current['phone']
                              else ('NEW' if not current['phone'] else 'CONFLICT'),
                    'source': source_label,
                })
            if sig.get('title'):
                findings.append({
                    'field': 'title',
                    'label': 'Title',
                    'found': sig['title'],
                    'current': current['title'],
                    'status': 'CONFIRMED' if sig['title'] == current['title']
                              else ('NEW' if not current['title'] else 'CONFLICT'),
                    'source': 'email footer',
                })

        # Outlook contacts lookup
        outlook = lookup_outlook_contact(token, person['name'], person.get('email', ''))
        if outlook:
            for field_key, label, outlook_key in [
                ('phone', 'Phone', 'phone'),
                ('title', 'Title', 'title'),
                ('email', 'Email', 'email'),
            ]:
                # Skip if already found from email footer
                already_found = any(f['field'] == field_key for f in findings)
                val = outlook.get(outlook_key, '')
                if val and not already_found:
                    findings.append({
                        'field': field_key,
                        'label': label,
                        'found': val,
                        'current': current[field_key],
                        'status': 'CONFIRMED' if val == current[field_key]
                                  else ('NEW' if not current[field_key] else 'CONFLICT'),
                        'source': 'Outlook contacts',
                    })

    except Exception as e:
        graph_error = str(e)
        print(f"[enrich] Graph API error for {slug}: {e}")

    enriched_at = datetime.now().isoformat()
    return jsonify({
        'findings': findings,
        'enriched_at': enriched_at,
        'graph_error': graph_error,
        'no_graph_consent': graph_error is not None,
    })


@crm_bp.route('/api/people/<slug>/enrich/save', methods=['POST'])
@login_required
def api_person_enrich_save(slug):
    """
    Save confirmed enrichment results to the contact record.
    Expects: { fields: {field: value, ...}, sources: {field: source, ...} }
    """
    person = load_person(slug)
    if not person:
        return jsonify({'error': 'Person not found'}), 404

    data = request.get_json(force=True)
    incoming = data.get('fields', {})
    sources = data.get('sources', {})

    save_fields = {}
    for key in ('phone', 'title', 'email', 'linkedin_url'):
        if key in incoming and incoming[key]:
            save_fields[key] = incoming[key]

    if sources:
        save_fields['enrichment_source'] = sources

    ok = save_enrichment_results(person['name'], save_fields)
    if not ok:
        return jsonify({'error': 'Save failed — contact not found in DB'}), 500

    enriched_at = datetime.now().isoformat()
    return jsonify({'ok': True, 'enriched_at': enriched_at})


# ---------------------------------------------------------------------------
# Offerings / Prospects API
# ---------------------------------------------------------------------------


@crm_bp.route('/api/prospects')
@login_required
def api_prospects():
    offering = request.args.get('offering', '')
    include_closed = request.args.get('include_closed', 'false').lower() == 'true'
    prospects = load_prospects(offering if offering else None)
    if not include_closed:
        excluded = {'8. Closed', '0. Not Pursuing', '0. Declined'}
        prospects = [p for p in prospects if p.get('Stage', '') not in excluded]

    # Load organizations for Type enrichment
    orgs = {o['name']: o for o in load_organizations()}

    # Apply type filter if present
    type_filter = request.args.get('type')
    if type_filter:
        prospects = [p for p in prospects if orgs.get(p.get('org', ''), {}).get('Type') == type_filter]

    # Get tasks from PostgreSQL for each prospect
    for p in prospects:
        org_name = p.get('org', '')
        prospect_tasks = get_tasks_for_prospect_db(org_name)
        p['Tasks'] = ''  # No legacy markdown tasks
        p['_tasks'] = []  # Legacy field, empty
        p['prospect_tasks'] = prospect_tasks
        p['open_task_count'] = len(prospect_tasks)
        # Inject Type from org; Primary Contact is already on the prospect record
        org_data = orgs.get(org_name, {})
        p['Type'] = org_data.get('Type', '')

    all_briefs = load_all_briefs()
    prospect_briefs = all_briefs.get('prospect', {})
    for p in prospects:
        brief_key = f"{p.get('org', '')}::{p.get('offering', '')}"
        brief = prospect_briefs.get(brief_key, {})
        p['at_a_glance'] = brief.get('at_a_glance', '')
    return jsonify(prospects)


@crm_bp.route('/api/fund-summary')
@login_required
def api_fund_summary():
    offering = request.args.get('offering', '')
    if offering:
        return jsonify(get_fund_summary(offering))
    return jsonify(get_fund_summary_all())


# ---------------------------------------------------------------------------
# Inline field edit
# ---------------------------------------------------------------------------

@crm_bp.route('/api/prospect/field', methods=['PATCH'])
@login_required
def api_patch_prospect_field():
    data = request.get_json(force=True)
    org = data.get('org', '').strip()
    offering = data.get('offering', '').strip()
    field = data.get('field', '').strip().lower()
    raw_value = data.get('value', '')
    value = str(raw_value).strip() if isinstance(raw_value, str) else raw_value
    if not org or not offering or not field:
        return jsonify({'error': 'org, offering, and field are required'}), 400
    if field not in EDITABLE_FIELDS:
        return jsonify({'error': f'Field "{field}" is not editable'}), 400
    config = load_crm_config()
    if field == 'stage' and value not in config['stages'] and value != '':
        return jsonify({'error': f'Invalid stage: {value}'}), 400
    if field == 'assigned_to' and value != '':
        valid_names = {member['name'] for member in config['team']}
        valid_names.update(m['short'] for m in config.get('team_map', []))
        valid_names.update(m['full'] for m in config.get('team_map', []))
        if value not in valid_names:
            return jsonify({'error': f'Invalid team member: {value}'}), 400
    if field == 'closing' and value not in config['closing_options'] and value != '':
        return jsonify({'error': f'Invalid closing option: {value}'}), 400
    if field == 'target':
        parsed = _parse_currency(value)
        value = f"${parsed:,.0f}" if parsed else '$0'
    update_prospect_field(org, offering, field, value)
    updated = get_prospect(org, offering)
    return jsonify(updated)


# ---------------------------------------------------------------------------
# Tasks Page
# ---------------------------------------------------------------------------

PRIORITY_ORDER = {'Hi': 1, 'High': 1, 'Med': 2, 'Medium': 2, 'Lo': 3, 'Low': 3}


@crm_bp.route('/tasks')
@login_required
def crm_tasks():
    all_tasks = get_all_prospect_tasks_with_index()

    # Load prospect data to enrich tasks with target and offering
    prospects = load_prospects()
    org_prospect_map = {}
    for prospect in prospects:
        org = prospect.get('org', '')
        if org:
            org_prospect_map[org.lower()] = prospect

    # Enrich tasks with prospect data
    enriched_tasks = []
    for task in all_tasks:
        if not task.get('owner', '').strip():
            continue

        org = task.get('org', '')
        prospect = org_prospect_map.get(org.lower())
        if prospect:
            target_str = prospect.get('Target', '$0')
            target_val = _parse_currency(target_str)
            task['target'] = target_val
            task['target_display'] = target_str
            task['offering'] = prospect.get('offering', '')
        else:
            task['target'] = 0
            task['target_display'] = '—'
            task['offering'] = ''

        enriched_tasks.append(task)

    user_display = (g.user.get('display_name') or '').strip().lower()
    user_email = (g.user.get('email') or '').strip().lower()
    user_ids = {user_display, user_email} - {''}

    def sort_key(t):
        pri = PRIORITY_ORDER.get(t['priority'], 4)
        return (pri, -(t['target'] or 0))

    def add_url(tasks):
        result = []
        for t in tasks:
            offering = t.get('offering', '')
            org = t.get('org', '')
            if offering and org:
                url = f"/crm/prospect/{urlquote(offering, safe='')}/{urlquote(org, safe='')}/detail"
            else:
                url = '#'
            result.append({**t, 'detail_url': url})
        return result

    my_tasks = add_url(sorted(
        [t for t in enriched_tasks if t['owner'].strip().lower() in user_ids],
        key=sort_key
    ))
    team_tasks = add_url(sorted(
        [t for t in enriched_tasks if t['owner'].strip().lower() not in user_ids],
        key=sort_key
    ))

    # Group by prospect (org)
    from collections import OrderedDict
    by_prospect = OrderedDict()
    for t in sorted(enriched_tasks, key=sort_key):
        org = t.get('org', '') or 'Untagged'
        by_prospect.setdefault(org, []).append(t)
    groups_by_prospect = [{'org': org, 'tasks': add_url(tasks)} for org, tasks in by_prospect.items()]

    # Group by owner
    by_owner = OrderedDict()
    for t in sorted(enriched_tasks, key=sort_key):
        owner = t.get('owner', '').strip() or 'Unassigned'
        by_owner.setdefault(owner, []).append(t)
    groups_by_owner = [{'owner': owner, 'tasks': add_url(tasks)} for owner, tasks in by_owner.items()]

    # All prospect org names for the add-task dropdown
    all_prospect_orgs = sorted({t.get('org', '') for t in enriched_tasks if t.get('org', '').strip()})

    active_view = request.args.get('view', 'prospect')
    config = load_crm_config()

    return render_template(
        'crm_tasks.html',
        active_tab='tasks',
        my_tasks=my_tasks,
        team_tasks=team_tasks,
        groups_by_prospect=groups_by_prospect,
        groups_by_owner=groups_by_owner,
        total_tasks=len(enriched_tasks),
        all_prospect_orgs=all_prospect_orgs,
        config=config,
        active_view=active_view,
    )


# ---------------------------------------------------------------------------
# Prospect Task API
# ---------------------------------------------------------------------------

@crm_bp.route('/api/tasks', methods=['GET'])
@login_required
def api_crm_tasks_list():
    org = request.args.get('org', '').strip()
    if not org:
        return jsonify({'error': 'org parameter required'}), 400
    raw_tasks = get_tasks_for_prospect_db(org)
    # Normalize field names to match the frontend contract
    # (task-edit-modal.js and crm_prospect_detail.html expect these keys)
    STATUS_MAP = {'open': 'New', 'done': 'Complete'}
    results = []
    for t in raw_tasks:
        results.append({
            'text': t.get('text', ''),
            'assigned_to': t.get('owner', ''),
            'priority': t.get('priority', 'Med'),
            'status': STATUS_MAP.get(t.get('status', ''), t.get('status', 'New')),
            'section': t.get('section', 'Fundraising - Me'),
            'index': t.get('section_index', 0),
            'context': '',
            'org': org,
        })
    return jsonify(results)


@crm_bp.route('/api/tasks', methods=['POST'])
@login_required
def api_crm_tasks_create():
    data = request.get_json(force=True)
    org = data.get('org', '').strip()
    text = data.get('text', '').strip()
    owner = data.get('owner', '').strip()
    priority = data.get('priority', 'Med').strip()
    if not org or not text or not owner:
        return jsonify({'ok': False, 'error': 'org, text, and owner are required'}), 400
    success = add_prospect_task(org, text, owner, priority)
    if not success:
        return jsonify({'ok': False, 'error': 'Failed to add task'}), 500
    # Return the newly created task by fetching it
    tasks = get_tasks_for_prospect_db(org)
    # Find the most recently added task matching this text
    for task in reversed(tasks):
        if task.get('text') == text:
            return jsonify({'ok': True, **task}), 201
    return jsonify({'ok': False, 'error': 'Task created but could not retrieve'}), 500


@crm_bp.route('/api/tasks/complete', methods=['PATCH'])
@login_required
def api_crm_tasks_complete():
    data = request.get_json(force=True)
    org = data.get('org', '').strip()
    text = data.get('text', '').strip()
    if not org or not text:
        return jsonify({'ok': False, 'error': 'org and text are required'}), 400
    success = complete_prospect_task(org, text)
    if not success:
        return jsonify({'ok': False, 'error': 'Task not found'}), 404
    return jsonify({'ok': True})


@crm_bp.route('/api/tasks/<int:task_id>', methods=['PATCH'])
@login_required
def api_crm_tasks_update(task_id):
    """Update task fields: text, owner, priority, status.

    NOTE: This endpoint is not yet implemented for markdown-backed tasks.
    Task updates must be done by editing TASKS.md directly.
    """
    return jsonify({'ok': False, 'error': 'Task update not implemented for markdown backend'}), 501


# ---------------------------------------------------------------------------
# Org API
# ---------------------------------------------------------------------------

@crm_bp.route('/api/org/<path:name>', methods=['GET'])
@login_required
def api_org_get(name):
    org = get_organization(name)
    if not org:
        org = {'name': name, 'Type': '', 'Notes': ''}
    contacts = get_contacts_for_org(name)
    prospects = get_prospects_for_org(name)
    saved = load_saved_brief('org', name)
    return jsonify({
        'org': org,
        'contacts': contacts,
        'prospects': prospects,
        'saved_brief': saved,
    })


@crm_bp.route('/api/org/<path:name>', methods=['PATCH'])
@login_required
def api_org_patch(name):
    data = request.get_json(force=True)
    payload = {}
    if 'type' in data:
        payload['Type'] = data['type']
    if 'domain' in data:
        payload['Domain'] = data['domain']
    if 'aliases' in data:
        payload['Aliases'] = data['aliases']
    if 'notes' in data:
        payload['Notes'] = data['notes']
    if not payload:
        return jsonify({'error': 'No valid fields to update'}), 400
    existing = get_organization(name)
    if existing:
        merged = {**existing, **payload}
    else:
        merged = {'name': name, 'Type': '', 'Notes': '', **payload}
    write_organization(name, merged)
    return jsonify(get_organization(name) or merged)


@crm_bp.route('/api/synthesize-org-brief', methods=['POST'])
@login_required
def api_synthesize_org_brief():
    """Synthesize an AI relationship brief for an organization."""
    data = request.get_json(force=True)
    org_name = data.get('org', '').strip()
    if not org_name:
        return jsonify({'error': 'org required'}), 400
    org = get_organization(org_name) or {'name': org_name}
    contacts = get_contacts_for_org(org_name)
    prospects = get_prospects_for_org(org_name)
    interactions = load_interactions(org=org_name, limit=10)
    emails = get_emails_for_org(org_name)[:10]
    context_lines = [f"Organization: {org_name}"]
    if org.get('Type'):
        context_lines.append(f"Type: {org['Type']}")
    if org.get('Domain'):
        context_lines.append(f"Domain: {org['Domain']}")
    if org.get('Notes'):
        context_lines.append(f"Notes: {org['Notes']}")
    if contacts:
        context_lines.append(f"\nContacts ({len(contacts)}):")
        for c in contacts:
            context_lines.append(f"  - {c.get('name', '')} ({c.get('role', '')})")
    if prospects:
        context_lines.append(f"\nProspects ({len(prospects)}):")
        for p in prospects:
            context_lines.append(
                f"  - {p.get('offering', '')} | Stage: {p.get('stage', '')} | "
                f"Target: {p.get('target', '')}"
            )
    if interactions:
        context_lines.append(f"\nRecent Interactions ({len(interactions)}):")
        for i in interactions[:5]:
            context_lines.append(f"  - [{i.get('date', '')}] {i.get('summary', '')}")
    if emails:
        context_lines.append(f"\nRecent Emails ({len(emails)}):")
        for e in emails[:5]:
            context_lines.append(
                f"  - [{e.get('date', '')}] {e.get('subject', '')} — {e.get('summary', '')}"
            )
    context_block = "\n".join(context_lines)
    content_hash = hashlib.md5(context_block.encode()).hexdigest()[:12]
    try:
        narrative, _ = call_claude_brief(
            ORG_BRIEF_SYSTEM,
            f"Generate an organizational brief for {org_name}.\n\n{context_block}",
            max_tokens=800,
            want_json=False,
        )
    except Exception as e:
        print(f"[org-brief] synthesis failed for {org_name}: {e}")
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e), 'narrative': ''}), 500
    save_brief('org', org_name, narrative, content_hash)
    return jsonify({
        'narrative': narrative,
        'content_hash': content_hash,
    })


@crm_bp.route('/api/org/<path:name>/notes', methods=['GET'])
@login_required
def api_org_notes_get(name):
    """Get org-level notes."""
    notes = load_org_notes(name)
    return jsonify(notes)


@crm_bp.route('/api/org/<path:name>/notes', methods=['POST'])
@login_required
def api_org_notes_post(name):
    """Add a new org-level note."""
    name = resolve_org_name(name)
    data = request.get_json(force=True)
    author = (g.user.get('display_name') or g.user.get('email') or 'Unknown').strip()
    text = (data.get('text') or '').strip()
    if not text:
        return jsonify({'error': 'text is required'}), 400
    entry = save_org_note(name, author, text)
    all_notes = load_org_notes(name)
    return jsonify({'ok': True, 'entry': entry, 'notes_log': all_notes})


# ---------------------------------------------------------------------------
# Contact API
# ---------------------------------------------------------------------------

@crm_bp.route('/api/org/<path:org_name>/contacts', methods=['POST'])
@login_required
def api_org_add_contact(org_name):
    """Add a contact to an organization."""
    org_name = resolve_org_name(org_name)
    data = request.get_json(force=True)
    name = data.get('name', '').strip()
    title = data.get('title', '').strip()
    email = data.get('email', '').strip()
    phone = data.get('phone', '').strip()
    is_new = data.get('is_new', False)

    if not name:
        return jsonify({'error': 'name is required'}), 400

    # If linking existing person, just update their org field
    # If creating new, create a person file
    if is_new:
        person_type = 'investor'
        slug = create_person_file(name, org_name, email, title, person_type)
        if phone:
            # Update phone field in the newly created file
            update_contact_fields(org_name, name, {'phone': phone})
    else:
        # Link existing person to this org
        people_dir = os.path.join(PROJECT_ROOT, 'contacts')
        # Find the person file by name
        found = False
        for fname in os.listdir(people_dir):
            if not fname.endswith('.md'):
                continue
            path = os.path.join(people_dir, fname)
            person = parse_kb_person_file(path)
            if person['name'].lower() == name.lower():
                # Update organization field
                slug = person['slug']
                update_contact_fields(org_name, name, {
                    'company': org_name,
                    'title': title or person.get('title', ''),
                    'email': email or person.get('email', ''),
                    'phone': phone or person.get('phone', ''),
                })
                found = True
                break
        if not found:
            return jsonify({'error': 'Person not found'}), 404

    # Auto-set as primary if this is the org's first contact
    all_contacts = get_contacts_for_org(org_name)
    if len(all_contacts) == 1:
        set_primary_contact(org_name, name)

    person = load_person(slug)
    return jsonify({'ok': True, 'person': person}), 201


@crm_bp.route('/api/org/<path:org_name>/primary-contact', methods=['POST'])
@login_required
def api_org_set_primary_contact(org_name):
    """Set or clear the primary contact for an org.
    Body: {"contact_name": "Julia McArdle"} to set, {"contact_name": null} to clear.
    """
    data = request.get_json(force=True)
    contact_name = data.get('contact_name')
    if not contact_name:
        clear_primary_contact(org_name)
        return jsonify({'status': 'ok', 'primary_contact': None})
    success = set_primary_contact(org_name, contact_name)
    if not success:
        return jsonify({'error': 'contact not found'}), 404
    return jsonify({'status': 'ok', 'primary_contact': contact_name})


@crm_bp.route('/api/contact', methods=['POST'])
@login_required
def api_contact_create():
    data = request.get_json(force=True)
    name = data.get('name', '').strip()
    org = data.get('org', '').strip()
    if not name or not org:
        return jsonify({'error': 'name and org are required'}), 400
    email = data.get('email', '').strip()
    role = data.get('role', '').strip()
    person_type = data.get('type', 'investor').strip()
    slug = create_person_file(name, org, email, role, person_type)
    person = load_person(slug)
    return jsonify(person), 201


@crm_bp.route('/api/contact/<path:org_and_name>', methods=['PATCH'])
@login_required
def api_contact_patch(org_and_name):
    parts = org_and_name.rsplit('/', 1)
    if len(parts) != 2:
        return jsonify({'error': 'URL must be /api/contact/<org>/<name>'}), 400
    org, name = parts[0], parts[1]
    data = request.get_json(force=True)
    allowed_fields = {'role', 'email', 'phone', 'title'}
    payload = {k: v for k, v in data.items() if k in allowed_fields}
    if not payload:
        return jsonify({'error': 'No valid fields to update'}), 400
    success = update_contact_fields(org, name, payload)
    if not success:
        return jsonify({'error': 'Contact not found'}), 404
    return jsonify({'ok': True})


# ---------------------------------------------------------------------------
# Prospect save / create / delete
# ---------------------------------------------------------------------------

@crm_bp.route('/api/prospect/save', methods=['POST'])
@login_required
def api_prospect_save():
    data = request.get_json(force=True)
    org = data.get('org', '').strip()
    offering = data.get('offering', '').strip()
    fields = data.get('fields', {})
    if not org or not offering:
        return jsonify({'error': 'org and offering are required'}), 400
    if not get_prospect(org, offering):
        return jsonify({'error': 'Prospect not found'}), 404
    for field, value in fields.items():
        update_prospect_field(org, offering, field, str(value))
    update_prospect_field(org, offering, 'last_touch', date.today().isoformat())
    return jsonify({'status': 'ok'})


@crm_bp.route('/api/prospect', methods=['POST'])
@login_required
def api_prospect_create():
    data = request.get_json(force=True)
    org = data.get('org', '').strip()
    org = resolve_org_name(org) if org else ''
    offering = data.get('offering', '').strip()
    if not org or not offering:
        return jsonify({'error': 'org and offering are required'}), 400
    existing = get_prospect(org, offering)
    if existing:
        return jsonify({'error': 'Prospect already exists for this org + offering'}), 409
    new_prospect = {
        'org': org,
        'offering': offering,
        'Stage': data.get('stage', '1. Prospect'),
        'Target': data.get('target', '$0'),
        'Committed': '$0',
        'Closing': '',
        'Urgent': '',
        'Assigned To': '',
        'Notes': '',
        'Last Touch': '',
    }
    write_prospect(org, offering, new_prospect)
    created = get_prospect(org, offering)
    return jsonify(created), 201


@crm_bp.route('/api/prospect', methods=['DELETE'])
@login_required
def api_prospect_delete():
    data = request.get_json(force=True)
    org = data.get('org', '').strip()
    offering = data.get('offering', '').strip()
    if not org or not offering:
        return jsonify({'error': 'org and offering required'}), 400
    try:
        delete_prospect(org, offering)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ---------------------------------------------------------------------------
# Unmatched review
# ---------------------------------------------------------------------------

@crm_bp.route('/api/unmatched', methods=['GET'])
@login_required
def api_unmatched_list():
    return jsonify(load_unmatched())


@crm_bp.route('/api/unmatched/resolve', methods=['POST'])
@login_required
def api_unmatched_resolve():
    data = request.get_json(force=True)
    email = data.get('participant_email', '').strip()
    org = data.get('org', '').strip()
    offering = data.get('offering', '').strip()
    int_type = data.get('type', 'Email').strip()
    subject = data.get('subject', '').strip()
    int_date = data.get('date', date.today().isoformat())
    if not email or not org:
        return jsonify({'error': 'participant_email and org required'}), 400
    append_interaction({
        'org': org,
        'type': int_type,
        'offering': offering,
        'date': int_date,
        'contact': data.get('participant_name', ''),
        'subject': subject,
        'summary': f'Manual resolve: {subject}',
        'source': 'manual',
    })
    remove_unmatched(email)
    return jsonify({'ok': True})


@crm_bp.route('/api/unmatched/<path:email>', methods=['DELETE'])
@login_required
def api_unmatched_dismiss(email):
    remove_unmatched(email)
    return jsonify({'ok': True})


# ---------------------------------------------------------------------------
# Auto-capture / Org list / Export / Org create
# ---------------------------------------------------------------------------

@crm_bp.route('/api/auto-capture', methods=['POST'])
@login_required
def api_auto_capture():
    try:
        from auth.graph_auth import get_access_token
        from sources.crm_graph_sync import run_auto_capture
        token = get_access_token(allow_device_flow=False)
        stats = run_auto_capture(token)
        return jsonify({
            'ok': True,
            'emails_scanned': stats.get('matched', 0) + stats.get('unmatched', 0) + stats.get('skipped_dedup', 0),
            'meetings_scanned': 0,
            'interactions_logged': stats.get('matched', 0),
            'prospects_touched': stats.get('matched', 0),
            'duplicates_skipped': stats.get('skipped_dedup', 0),
            'unmatched_count': stats.get('unmatched', 0),
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@crm_bp.route('/api/orgs')
@login_required
def api_orgs():
    orgs = load_organizations()
    return jsonify([o['name'] for o in orgs])


@crm_bp.route('/api/export')
@login_required
def api_export_pipeline():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    offering = request.args.get('offering')
    if not offering:
        return jsonify({"error": "offering required"}), 400

    prospects = load_prospects(offering)
    orgs = {o['name']: o for o in load_organizations()}

    stage_filter = request.args.get('stage')
    urgent_only = request.args.get('urgent') == 'true'
    type_filter = request.args.get('type')
    assigned_filter = request.args.get('assigned')

    if stage_filter:
        prospects = [p for p in prospects if p.get('Stage') == stage_filter]
    if urgent_only:
        prospects = [p for p in prospects if p.get('Urgent')]
    if type_filter:
        prospects = [p for p in prospects if orgs.get(p.get('org', ''), {}).get('Type') == type_filter]
    if assigned_filter:
        prospects = [p for p in prospects if assigned_filter in str(p.get('Assigned To', ''))]

    def sort_key(p):
        stage = p.get('Stage', '0. Unknown')
        stage_num = 0
        if stage and stage[0].isdigit():
            try:
                stage_num = int(stage.split('.')[0])
            except (ValueError, IndexError):
                stage_num = 0
        urgent = str(p.get('Urgent', '')).lower()
        urgency_order = {'yes': 0, 'high': 0, 'med': 1, 'medium': 1, 'low': 2, '': 3}
        urgency_val = urgency_order.get(urgent, 3)
        target = p.get('Target', '$0')
        target_val = _parse_currency(target) if target else 0
        return (-stage_num, urgency_val, -target_val)

    prospects.sort(key=sort_key)

    wb = Workbook()
    ws = wb.active
    ws.title = "Pipeline"

    columns = [
        ('Organization', 30), ('Type', 16), ('Stage', 20), ('Target', 16),
        ('Committed', 16), ('Closing', 10), ('Urgency', 10), ('Assigned To', 20),
        ('Primary Contact', 22), ('Notes', 40), ('Last Touch', 14),
    ]

    for idx, (col_name, width) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    header_fill = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
    header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center')

    for idx, (col_name, _) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
    ws.freeze_panes = 'A2'

    row_fill_alt = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
    data_font = Font(name='Arial', size=10)
    today = date.today()

    for row_idx, p in enumerate(prospects, start=2):
        org_name = p.get('org', '')
        org_data = orgs.get(org_name, {})
        org_type = org_data.get('Type', '')
        if row_idx % 2 == 0:
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = row_fill_alt

        cell_org = ws.cell(row=row_idx, column=1, value=org_name)
        cell_org.font = Font(name='Arial', size=10, bold=True)
        cell_org.alignment = Alignment(horizontal='left', vertical='top')

        cell_type = ws.cell(row=row_idx, column=2, value=org_type)
        cell_type.font = data_font
        cell_type.alignment = Alignment(horizontal='left', vertical='top')

        stage = p.get('Stage', '')
        cell_stage = ws.cell(row=row_idx, column=3, value=stage)
        cell_stage.font = data_font
        cell_stage.alignment = Alignment(horizontal='left', vertical='top')

        target_str = p.get('Target', '$0')
        target_val = _parse_currency(target_str) if target_str else 0
        cell_target = ws.cell(row=row_idx, column=4, value=target_val)
        cell_target.font = data_font
        cell_target.number_format = '$#,##0'
        cell_target.alignment = Alignment(horizontal='right', vertical='top')

        committed_str = p.get('Committed', '$0')
        committed_val = _parse_currency(committed_str) if committed_str else 0
        cell_committed = ws.cell(row=row_idx, column=5, value=committed_val)
        cell_committed.font = data_font
        cell_committed.number_format = '$#,##0'
        cell_committed.alignment = Alignment(horizontal='right', vertical='top')

        closing = p.get('Closing', '')
        cell_closing = ws.cell(row=row_idx, column=6, value=closing)
        cell_closing.font = data_font
        cell_closing.alignment = Alignment(horizontal='center', vertical='top')

        urgency = p.get('Urgency', '').strip()
        cell_urgency = ws.cell(row=row_idx, column=7, value=urgency)
        cell_urgency.font = data_font
        cell_urgency.alignment = Alignment(horizontal='center', vertical='top')

        if urgency == 'High':
            cell_urgency.fill = PatternFill(start_color='fef2f2', end_color='fef2f2', fill_type='solid')
            cell_urgency.font = Font(name='Arial', size=10, color='ef4444')
        elif urgency == 'Med':
            cell_urgency.fill = PatternFill(start_color='fffbeb', end_color='fffbeb', fill_type='solid')
            cell_urgency.font = Font(name='Arial', size=10, color='f59e0b')
        elif urgency == 'Low':
            cell_urgency.fill = PatternFill(start_color='f9fafb', end_color='f9fafb', fill_type='solid')
            cell_urgency.font = Font(name='Arial', size=10, color='9ca3af')

        assigned = p.get('Assigned To', '')
        cell_assigned = ws.cell(row=row_idx, column=8, value=assigned)
        cell_assigned.font = data_font
        cell_assigned.alignment = Alignment(horizontal='left', vertical='top')

        primary_contact = p.get('Primary Contact', '')
        cell_contact = ws.cell(row=row_idx, column=9, value=primary_contact)
        cell_contact.font = data_font
        cell_contact.alignment = Alignment(horizontal='left', vertical='top')

        notes = p.get('Notes', '')
        cell_notes = ws.cell(row=row_idx, column=10, value=notes)
        cell_notes.font = data_font
        cell_notes.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        last_touch = p.get('Last Touch', '')
        cell_last_touch = ws.cell(row=row_idx, column=11, value=last_touch)
        cell_last_touch.font = data_font
        cell_last_touch.alignment = Alignment(horizontal='center', vertical='top')

        if last_touch:
            try:
                touch_date = datetime.strptime(last_touch, '%Y-%m-%d').date()
                days_ago = (today - touch_date).days
                if days_ago < 7:
                    cell_last_touch.fill = PatternFill(start_color='f0fdf4', end_color='f0fdf4', fill_type='solid')
                elif days_ago <= 14:
                    cell_last_touch.fill = PatternFill(start_color='fffbeb', end_color='fffbeb', fill_type='solid')
                else:
                    cell_last_touch.fill = PatternFill(start_color='fef2f2', end_color='fef2f2', fill_type='solid')
            except (ValueError, AttributeError):
                pass

    summary_row = len(prospects) + 2
    cell_total_label = ws.cell(row=summary_row, column=1, value=f"TOTAL ({len(prospects)} prospects)")
    cell_total_label.font = Font(name='Arial', size=10, bold=True)

    cell_target_sum = ws.cell(row=summary_row, column=4, value=f"=SUM(D2:D{summary_row-1})")
    cell_target_sum.font = Font(name='Arial', size=10, bold=True)
    cell_target_sum.number_format = '$#,##0'
    cell_target_sum.alignment = Alignment(horizontal='right')

    cell_committed_sum = ws.cell(row=summary_row, column=5, value=f"=SUM(E2:E{summary_row-1})")
    cell_committed_sum.font = Font(name='Arial', size=10, bold=True)
    cell_committed_sum.number_format = '$#,##0'
    cell_committed_sum.alignment = Alignment(horizontal='right')

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_offering = offering.replace(' ', '_').replace('/', '_')
    filename = f"AREC_Pipeline_{safe_offering}_{today.isoformat()}.xlsx"

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@crm_bp.route('/api/org', methods=['POST'])
@login_required
def api_org_create():
    data = request.get_json(force=True)
    name = data.get('name', '').strip()
    org_type = data.get('type', '').strip()
    domain = data.get('domain', '').strip()
    notes = data.get('notes', '').strip()
    if not name:
        return jsonify({'error': 'Organization name is required'}), 400
    existing = get_organization(name)
    if existing:
        return jsonify({'error': 'Organization already exists'}), 409
    org_data = {'name': name, 'Type': org_type, 'Domain': domain, 'Notes': notes}
    write_organization(name, org_data)
    return jsonify({'ok': True, 'org': org_data}), 201


# ---------------------------------------------------------------------------
# Meeting History
# ---------------------------------------------------------------------------

@crm_bp.route('/api/org/<path:name>/meetings', methods=['GET'])
@login_required
def api_org_meetings(name):
    meetings = load_meeting_history(name)
    return jsonify(meetings)


@crm_bp.route('/api/org/<path:name>/meetings', methods=['POST'])
@login_required
def api_org_meeting_add(name):
    data = request.get_json(force=True)
    add_meeting_entry(
        org=name,
        date=data.get('date', ''),
        title=data.get('title', ''),
        attendees=data.get('attendees', ''),
        source=data.get('source', 'manual'),
        notion_url=data.get('notion_url', ''),
    )
    return jsonify({'ok': True})


@crm_bp.route('/api/org/<path:source>/merge-preview', methods=['GET'])
@login_required
def api_org_merge_preview(source):
    """Preview what will be merged when merging source into target."""
    target = request.args.get('target', '').strip()
    if not target:
        return jsonify({'error': 'target parameter required'}), 400

    try:
        preview = get_merge_preview(source, target)
        return jsonify(preview)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@crm_bp.route('/api/org/merge', methods=['POST'])
@login_required
def api_org_merge():
    """Merge source org into target org."""
    data = request.get_json(force=True)
    source = data.get('source', '').strip()
    target = data.get('target', '').strip()

    if not source or not target:
        return jsonify({'error': 'source and target required'}), 400

    if source.lower() == target.lower():
        return jsonify({'error': 'Cannot merge an org into itself'}), 400

    try:
        result = merge_organizations(source, target)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# /api/followup endpoint removed — use /crm/api/tasks POST instead
# Tasks are now managed in prospect_tasks table, not TASKS.md


# ---------------------------------------------------------------------------
# Health / Engagement Heatmap
# ---------------------------------------------------------------------------

@crm_bp.route('/health')
@login_required
def health_page():
    from datetime import datetime as _dt
    prospects = get_heatmap_prospects()

    STATUS_ORDER = ['no_contact', 'outbound_only', 'inbound', 'held', 'scheduled']

    # Sort: worst status first, within tier by target descending
    def sort_key(p):
        status_rank = STATUS_ORDER.index(p['status']) if p['status'] in STATUS_ORDER else 0
        return (status_rank, -(p['target'] or 0))

    prospects.sort(key=sort_key)

    # Group by status
    from collections import OrderedDict
    groups = OrderedDict()
    for key in STATUS_ORDER:
        groups[key] = []
    for p in prospects:
        groups[p['status']].append(p)

    last_updated = _dt.now().strftime('%-I:%M %p')
    return render_template(
        'crm_health.html',
        active_tab='health',
        groups=groups,
        total=len(prospects),
        last_updated=last_updated,
    )


# ---------------------------------------------------------------------------
# Meetings Page and API
# ---------------------------------------------------------------------------

@crm_bp.route('/meetings')
@login_required
def meetings_page():
    return render_template('crm_meetings.html', active_tab='meetings')


@crm_bp.route('/api/offerings', methods=['GET'])
@login_required
def api_offerings_list():
    """Return list of offering names as JSON array."""
    offerings = load_offerings()
    return jsonify([o['name'] for o in offerings])


@crm_bp.route('/api/meetings', methods=['GET'])
@login_required
def api_meetings_list():
    """Return all meetings as JSON array."""
    meetings = load_meetings()
    return jsonify(meetings)


@crm_bp.route('/api/meetings', methods=['POST'])
@login_required
def api_meetings_create():
    """Create a new meeting."""
    data = request.get_json(force=True)
    org = data.get('org', '').strip()
    org = resolve_org_name(org) if org else ''
    offering = data.get('offering', '').strip()
    meeting_date = data.get('meeting_date', '').strip()

    if not meeting_date:
        return jsonify({'error': 'meeting_date is required'}), 400

    # Business rule: offering requires org
    if offering and not org:
        return jsonify({'error': 'Cannot set offering without organization'}), 400

    # Validate offering against known offerings if provided
    if offering:
        valid_offerings = [o['name'] for o in load_offerings()]
        if offering not in valid_offerings:
            return jsonify({'error': f'Invalid offering. Must be one of: {", ".join(valid_offerings)}'}), 400

    created_by = (g.user.get('email') or g.user.get('display_name', 'oscar')).strip()

    meeting = save_meeting(
        org=org,
        offering=offering,
        meeting_date=meeting_date,
        meeting_time=data.get('meeting_time', ''),
        title=data.get('title', ''),
        attendees=data.get('attendees', ''),
        source=data.get('source', 'manual'),
        graph_event_id=data.get('graph_event_id'),
        notes_raw=data.get('notes_raw', ''),
        transcript_url=data.get('transcript_url', ''),
        created_by=created_by,
    )

    return jsonify({'ok': True, 'meeting': meeting}), 201


@crm_bp.route('/api/meetings/<meeting_id>', methods=['GET'])
@login_required
def api_meetings_get(meeting_id):
    """Get single meeting by UUID."""
    meeting = get_meeting(meeting_id)
    if not meeting:
        abort(404)
    return jsonify(meeting)


@crm_bp.route('/api/meetings/<meeting_id>', methods=['PATCH'])
@login_required
def api_meetings_update(meeting_id):
    """Update fields on a meeting."""
    data = request.get_json(force=True)
    # Normalize org name if present
    if 'org' in data and data['org']:
        data['org'] = resolve_org_name(data['org'].strip())
    updated = update_meeting(meeting_id, **data)
    if not updated:
        abort(404)
    return jsonify(updated)


@crm_bp.route('/api/meetings/<meeting_id>', methods=['DELETE'])
@login_required
def api_meetings_delete(meeting_id):
    """Delete a meeting."""
    success = delete_meeting(meeting_id)
    if not success:
        abort(404)
    return jsonify({'ok': True})


@crm_bp.route('/api/meetings/<meeting_id>/notes', methods=['POST'])
@login_required
def api_meetings_notes(meeting_id):
    """Attach notes to a meeting and trigger AI processing."""
    data = request.get_json(force=True)
    notes_raw = data.get('notes_raw', '').strip()

    if not notes_raw:
        return jsonify({'error': 'notes_raw is required'}), 400

    # Update meeting with notes
    updated = update_meeting(meeting_id, notes_raw=notes_raw)
    if not updated:
        abort(404)

    # Process notes with AI
    processed = process_meeting_notes(meeting_id)
    return jsonify(processed)


@crm_bp.route('/api/meetings/<meeting_id>/insights/<insight_id>/approve', methods=['POST'])
@login_required
def api_meetings_insight_approve(meeting_id, insight_id):
    """Approve a meeting insight."""
    username = g.user.get('email') or g.user.get('display_name', 'oscar')
    updated = approve_meeting_insight(meeting_id, insight_id, username)
    if not updated:
        abort(404)
    return jsonify(updated)


@crm_bp.route('/api/meetings/<meeting_id>/insights/<insight_id>/dismiss', methods=['POST'])
@login_required
def api_meetings_insight_dismiss(meeting_id, insight_id):
    """Dismiss a meeting insight."""
    username = g.user.get('email') or g.user.get('display_name', 'oscar')
    updated = dismiss_meeting_insight(meeting_id, insight_id, username)
    if not updated:
        abort(404)
    return jsonify(updated)


# ---------------------------------------------------------------------------
# Flat Task CRUD — helpers (migrated from tasks_blueprint)
# ---------------------------------------------------------------------------

def _load_tasks_full() -> dict:
    """Parse TASKS.md → {'open': [...], 'done': [...]}. Each task includes 'index'."""
    if not os.path.exists(TASKS_PATH):
        return {'open': [], 'done': []}

    open_tasks = []
    done_tasks = []
    in_done = False

    with open(TASKS_PATH, encoding='utf-8') as f:
        for line in f:
            stripped = line.rstrip()
            if re.match(r'^## Done', stripped):
                in_done = True
                continue
            if re.match(r'^## ', stripped):
                continue
            if line.startswith('- [ ] ') or line.startswith('- [x] '):
                if in_done:
                    task = _parse_task_line(line)
                    task['index'] = len(done_tasks)
                    done_tasks.append(task)
                else:
                    task = _parse_task_line(line)
                    task['index'] = len(open_tasks)
                    open_tasks.append(task)

    return {'open': open_tasks, 'done': done_tasks}


def _read_task_lines() -> list:
    if not os.path.exists(TASKS_PATH):
        return []
    with open(TASKS_PATH, 'r', encoding='utf-8') as f:
        return f.readlines()


def _write_task_file(lines: list) -> None:
    with open(TASKS_PATH, 'w', encoding='utf-8') as f:
        f.writelines(lines)


def _find_task_line(lines: list, index: int, done: bool = False) -> int:
    """Find the file line number for the nth task (0-based).

    If done=False: counts open tasks before ## Done.
    If done=True: counts tasks after ## Done.
    """
    in_done_section = False
    count = 0
    for i, ln in enumerate(lines):
        if re.match(r'^## Done', ln.rstrip()):
            in_done_section = True
            continue
        if re.match(r'^## ', ln.rstrip()):
            in_done_section = False
            continue
        if in_done_section == done and (ln.startswith('- [ ] ') or ln.startswith('- [x] ')):
            if count == index:
                return i
            count += 1
    return -1


def _find_done_insertion_point(lines: list) -> int:
    """Return the line index immediately after '## Done' header."""
    for i, ln in enumerate(lines):
        if re.match(r'^## Done', ln.rstrip()):
            return i + 1
    return len(lines)


def _find_open_insertion_point(lines: list) -> int:
    """Return the line index just before '## Done' (to append open tasks)."""
    for i, ln in enumerate(lines):
        if re.match(r'^## Done', ln.rstrip()):
            return i
    return len(lines)


# ---------------------------------------------------------------------------
# Flat Task CRUD — routes (migrated from tasks_blueprint)
# ---------------------------------------------------------------------------

@crm_bp.route('/api/all-tasks', methods=['GET'])
@login_required
def api_all_tasks():
    """Return all tasks as {open: [...], done: [...]}."""
    return jsonify(_load_tasks_full())


@crm_bp.route('/api/task', methods=['POST'])
@login_required
def api_task_create():
    data = request.get_json(force=True)
    text = data.get('text', '').strip()
    priority = data.get('priority', 'Med').strip()
    context = data.get('context', '').strip()
    assigned_to = normalize_team_name(data.get('assigned_to', ''))
    status = data.get('status', 'new').strip()
    org = data.get('org', '').strip()

    if not text:
        return jsonify({'ok': False, 'error': 'text required'}), 400

    new_line = _format_task_line(text, priority, context, assigned_to, status=status, org=org)
    lines = _read_task_lines()
    insert_at = _find_open_insertion_point(lines)
    lines.insert(insert_at, new_line)
    _write_task_file(lines)
    return jsonify({'ok': True})


@crm_bp.route('/api/task/<int:index>', methods=['PUT'])
@login_required
def api_task_update(index):
    data = request.get_json(force=True)
    text = data.get('text', '').strip()
    priority = data.get('priority', 'Med').strip()
    context = data.get('context', '').strip()
    assigned_to = normalize_team_name(data.get('assigned_to', ''))
    status = data.get('status', 'open').strip()
    org = data.get('org', '').strip()

    if not text:
        return jsonify({'ok': False, 'error': 'text required'}), 400

    lines = _read_task_lines()
    li = _find_task_line(lines, index)
    if li == -1:
        return jsonify({'ok': False, 'error': 'task not found'}), 404

    original = lines[li]
    was_done = original.startswith('- [x] ')
    cdm = re.search(r'—\s*completed\s+(\d{4}-\d{2}-\d{2})', original)
    completion_date = cdm.group(1) if cdm else None

    new_line = _format_task_line(text, priority, context, assigned_to,
                                  done=was_done, completion_date=completion_date,
                                  status=status, org=org)
    lines[li] = new_line
    _write_task_file(lines)
    return jsonify({'ok': True})


@crm_bp.route('/api/task/<int:index>', methods=['DELETE'])
@login_required
def api_task_delete(index):
    lines = _read_task_lines()
    li = _find_task_line(lines, index)
    if li == -1:
        return jsonify({'ok': False, 'error': 'task not found'}), 404
    lines.pop(li)
    _write_task_file(lines)
    return jsonify({'ok': True})


@crm_bp.route('/api/task/<int:index>/complete', methods=['POST'])
@login_required
def api_task_complete(index):
    today = date.today().isoformat()
    lines = _read_task_lines()
    li = _find_task_line(lines, index)
    if li == -1:
        return jsonify({'ok': False, 'error': 'task not found'}), 404
    ln = lines[li]
    if not ln.startswith('- [ ] '):
        return jsonify({'ok': False, 'error': 'task already complete'}), 400

    ln = '- [x] ' + ln[6:].rstrip()
    if 'completed' not in ln:
        ln += f' — completed {today}'
    ln += '\n'

    lines.pop(li)
    insert_at = _find_done_insertion_point(lines)
    lines.insert(insert_at, ln)
    _write_task_file(lines)
    return jsonify({'ok': True})


@crm_bp.route('/api/task/<int:index>/restore', methods=['POST'])
@login_required
def api_task_restore(index):
    lines = _read_task_lines()
    li = _find_task_line(lines, index, done=True)
    if li == -1:
        return jsonify({'ok': False, 'error': 'task not found'}), 404
    ln = lines[li]
    if not ln.startswith('- [x] '):
        return jsonify({'ok': False, 'error': 'task not complete'}), 400

    ln = '- [ ] ' + ln[6:]
    ln = re.sub(r'\s*—\s*completed\s+\d{4}-\d{2}-\d{2}', '', ln.rstrip())
    ln += '\n'

    lines.pop(li)
    insert_at = _find_open_insertion_point(lines)
    lines.insert(insert_at, ln)
    _write_task_file(lines)
    return jsonify({'ok': True})


@crm_bp.route('/api/task/<int:index>/status', methods=['PATCH'])
@login_required
def api_task_status_update(index):
    data = request.get_json(force=True)
    new_status = data.get('status', '').strip()

    valid_statuses = ['New', 'In Progress', 'Complete']
    if new_status not in valid_statuses:
        return jsonify({'ok': False, 'error': f'status must be one of: {", ".join(valid_statuses)}'}), 400

    lines = _read_task_lines()
    li = _find_task_line(lines, index)
    if li == -1:
        return jsonify({'ok': False, 'error': 'task not found'}), 404

    task = _parse_task_line(lines[li])
    done = new_status == 'Complete'
    completion_date = date.today().isoformat() if done and not task['complete'] else task.get('completion_date')

    new_line = _format_task_line(
        task['text'], task['priority'], task['context'], task['assigned_to'],
        done=done, completion_date=completion_date,
        status=new_status, org=task.get('org', '')
    )
    lines[li] = new_line
    _write_task_file(lines)
    return jsonify({'ok': True, 'status': new_status})


@crm_bp.route('/api/task/<int:index>/priority', methods=['PATCH'])
@login_required
def api_task_priority_update(index):
    data = request.get_json(force=True)
    new_priority = data.get('priority', '').strip()

    valid_priorities = ['Hi', 'Med', 'Low']
    if new_priority not in valid_priorities:
        return jsonify({'ok': False, 'error': f'priority must be one of: {", ".join(valid_priorities)}'}), 400

    lines = _read_task_lines()
    li = _find_task_line(lines, index)
    if li == -1:
        return jsonify({'ok': False, 'error': 'task not found'}), 404

    task = _parse_task_line(lines[li])
    new_line = _format_task_line(
        task['text'], new_priority, task['context'], task['assigned_to'],
        done=task['complete'], completion_date=task.get('completion_date'),
        status=task.get('status', 'New'), org=task.get('org', '')
    )
    lines[li] = new_line
    _write_task_file(lines)
    return jsonify({'ok': True, 'priority': new_priority})
