"""
Flask app — productivity dashboard (port 3001) + CRM blueprint.
"""

import os
import sys
import re
import json
import hashlib
import glob as globmod
import anthropic

# Allow imports from app/
APP_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if APP_DIR not in sys.path:
    sys.path.insert(0, APP_DIR)

# Load .env from app/ directory
from dotenv import load_dotenv
load_dotenv(os.path.join(APP_DIR, ".env"))

from datetime import date, datetime, timezone, timedelta
from flask import Flask, Blueprint, jsonify, request, render_template, redirect, url_for, abort, send_file
from sources.crm_reader import (
    load_prospects, load_offerings, get_fund_summary, get_fund_summary_all,
    load_crm_config, get_organization, write_organization, load_organizations,
    get_contacts_for_org, create_person_file, update_contact_fields,
    get_prospects_for_org, get_prospect, write_prospect, update_prospect_field,
    load_unmatched, remove_unmatched, add_unmatched,
    _parse_currency, load_person, load_tasks_by_org, load_all_persons,
    delete_prospect, load_meeting_history, add_meeting_entry,
    get_tasks_for_prospect, get_all_prospect_tasks, add_prospect_task,
    complete_prospect_task,
    load_email_log, get_emails_for_org, find_email_by_message_id,
    load_interactions, append_interaction,
)
from sources.relationship_brief import (
    find_people_files, find_glossary_entry, find_meeting_summaries, find_org_tasks,
    collect_relationship_data, build_context_block, build_fallback_summary,
    compute_content_hash, BRIEF_SYSTEM_PROMPT,
    collect_person_data, build_person_context_block, build_person_fallback_summary,
    execute_person_updates, PERSON_BRIEF_SYSTEM_PROMPT, PERSON_UPDATE_ROUTING_PROMPT,
)

PROJECT_ROOT = os.path.dirname(APP_DIR)
TASKS_PATH = os.path.join(PROJECT_ROOT, "TASKS.md")

app = Flask(
    __name__,
    template_folder=os.path.join(APP_DIR, "templates"),
    static_folder=os.path.join(APP_DIR, "static"),
)

MEETINGS_DIR = os.path.join(PROJECT_ROOT, "meeting-summaries")
CALENDAR_PATH = os.path.join(PROJECT_ROOT, "dashboard_calendar.json")

# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

@app.route('/')
def dashboard():
    all_sections = _load_tasks_grouped()
    # Dashboard shows all sections (except Done) split across two columns
    tasks_by_section = [s for s in all_sections if s['name'] != 'Done']
    meetings = _load_recent_meetings(limit=10)
    calendar, calendar_stale = _load_calendar()
    config = load_crm_config()
    now = datetime.now()
    return render_template(
        'dashboard.html',
        tasks_by_section=tasks_by_section,
        meetings=meetings,
        calendar=calendar,
        calendar_stale=calendar_stale,
        config=config,
        now=now,
    )


def _load_tasks_grouped() -> list[dict]:
    """Read TASKS.md and return tasks grouped by section."""
    if not os.path.exists(TASKS_PATH):
        return []
    sections = []
    current = None
    with open(TASKS_PATH, encoding='utf-8') as f:
        for line in f:
            line = line.rstrip()
            m = re.match(r'^## (.+)$', line)
            if m:
                name = m.group(1).strip()
                current = {'name': name, 'tasks': []}
                sections.append(current)
                continue
            if current is not None and (line.startswith('- [ ] ') or line.startswith('- [x] ')):
                done = line.startswith('- [x] ')
                text = line[6:].strip()
                # Parse priority
                priority = 'Med'
                pm = re.match(r'\*\*\[(\w+)\]\*\*\s*', text)
                if pm:
                    priority = pm.group(1)
                    text = text[pm.end():]
                # Parse status (In Progress marker)
                status = 'Complete' if done else 'New'
                if '**[→]**' in text:
                    status = 'In Progress'
                    text = text.replace('**[→]**', '').strip()
                # Clean up old [STATUS:xxx] format for backward compatibility
                text = re.sub(r'\s*\[STATUS:\w+\]\s*', ' ', text).strip()
                # Extract assigned:Name inline field (new format)
                assigned_to = ''
                am = re.search(r'\s*—\s*assigned:([^—\n]+)', text)
                if am:
                    assigned_to = am.group(1).strip()
                    text = text[:am.start()] + text[am.end():]
                    text = text.strip()
                # Fallback: legacy **@Name** format
                if not assigned_to:
                    owner_m = re.match(r'\*\*@([^*]+)\*\*\s*', text)
                    if owner_m:
                        assigned_to = owner_m.group(1).strip()
                        text = text[owner_m.end():]
                # Extract (OrgName) suffix
                org = ''
                org_m = re.search(r'\(([^)]+)\)\s*$', text)
                if org_m:
                    org = org_m.group(1).strip()
                    text = text[:org_m.start()].rstrip(' —-')
                # Strip ~~...~~ for done items
                text = re.sub(r'~~(.+?)~~', r'\1', text)
                idx = len(current['tasks'])
                current['tasks'].append({
                    'text': text,
                    'done': done,
                    'status': status,
                    'priority': priority,
                    'org': org,
                    'assigned_to': assigned_to,
                    'raw': line[6:].strip(),
                    'index': idx,
                })
    # Sort tasks within each section by priority
    pri_order = {'Hi': 0, 'Med': 1, 'Low': 2}
    for sec in sections:
        sec['tasks'].sort(key=lambda t: pri_order.get(t['priority'], 1))
    return sections


def _load_tasks() -> list[dict]:
    """Read TASKS.md and return top-level task lines (flat, for API)."""
    if not os.path.exists(TASKS_PATH):
        return []
    tasks = []
    with open(TASKS_PATH, encoding='utf-8') as f:
        for line in f:
            line = line.rstrip()
            if line.startswith('- [ ] ') or line.startswith('- [x] '):
                done = line.startswith('- [x] ')
                tasks.append({'text': line[6:].strip(), 'done': done})
    return tasks[:50]


def _render_meeting_markdown(text: str) -> str:
    """Convert meeting summary markdown to safe HTML for display."""
    import html as html_mod
    lines = text.splitlines()
    out = []
    in_ul = False

    def _inline(s):
        s = html_mod.escape(s)
        s = re.sub(r'\*\*([^*\n]+)\*\*', r'<strong>\1</strong>', s)
        s = re.sub(r'\*([^*\n]+)\*', r'<em>\1</em>', s)
        s = re.sub(r'~~([^~\n]+)~~', r'<del>\1</del>', s)
        # Inline links
        s = re.sub(r'\[([^\]]+)\]\((https?://[^\)]+)\)',
                   r'<a href="\2" target="_blank" rel="noopener">\1</a>', s)
        return s

    def _close_ul():
        nonlocal in_ul
        if in_ul:
            out.append('</ul>')
            in_ul = False

    for line in lines:
        stripped = line.strip()

        # Skip raw header metadata lines (rendered in page header already)
        if stripped.startswith('**Date:**') or stripped.startswith('**Source:**') or stripped.startswith('**Attendees:**'):
            continue

        if stripped.startswith('# '):
            _close_ul()
            continue  # Title already shown in header
        elif stripped.startswith('## '):
            _close_ul()
            out.append(f'<h2>{html_mod.escape(stripped[3:])}</h2>')
        elif stripped.startswith('### '):
            _close_ul()
            out.append(f'<h3>{html_mod.escape(stripped[4:])}</h3>')
        elif stripped.startswith('- [x] ') or stripped.startswith('- [X] '):
            if not in_ul:
                out.append('<ul>'); in_ul = True
            out.append(f'<li class="checked">{_inline(stripped[6:])}</li>')
        elif stripped.startswith('- [ ] '):
            if not in_ul:
                out.append('<ul>'); in_ul = True
            out.append(f'<li>{_inline(stripped[6:])}</li>')
        elif stripped.startswith('- '):
            if not in_ul:
                out.append('<ul>'); in_ul = True
            out.append(f'<li>{_inline(stripped[2:])}</li>')
        elif re.match(r'^-{3,}$', stripped):
            _close_ul()
            out.append('<hr>')
        elif stripped == '':
            _close_ul()
        else:
            _close_ul()
            out.append(f'<p>{_inline(stripped)}</p>')

    _close_ul()
    return '\n'.join(out)


def _load_recent_meetings(limit=10) -> list[dict]:
    """Load recent meeting summaries from markdown files."""
    if not os.path.isdir(MEETINGS_DIR):
        return []
    files = sorted(globmod.glob(os.path.join(MEETINGS_DIR, "*.md")), reverse=True)
    meetings = []
    for fp in files[:limit]:
        fname = os.path.basename(fp)
        # Parse date and title from filename: YYYY-MM-DD-title-slug.md
        m = re.match(r'(\d{4}-\d{2}-\d{2})-(.+)\.md$', fname)
        if not m:
            continue
        meeting_date = m.group(1)
        slug = m.group(2)
        title = slug.replace('-', ' ').title()
        attendees = ''
        notion_url = ''
        with open(fp, encoding='utf-8') as f:
            for line in f:
                if line.startswith('**Attendees:**'):
                    attendees = line.split(':', 1)[1].strip()
                if line.startswith('**Source:**'):
                    um = re.search(r'\[.*?\]\((https?://[^\)]+)\)', line)
                    if um:
                        notion_url = um.group(1)
                if attendees and notion_url:
                    break
        meetings.append({
            'date': meeting_date,
            'title': title,
            'attendees': attendees,
            'url': notion_url,
            'filename': fname,
        })
    return meetings


def _load_calendar() -> tuple[list[dict], str | None]:
    """Load today's calendar from JSON file (written by update process).
    Returns (events, stale_date) where stale_date is set if the file is
    from a previous day (so the template can show a staleness warning).
    """
    if not os.path.exists(CALENDAR_PATH):
        return [], None
    try:
        mtime = os.path.getmtime(CALENDAR_PATH)
        file_date = datetime.fromtimestamp(mtime).date()
        stale = file_date if file_date != date.today() else None
        if stale:
            return [], stale.strftime('%b %-d')
        with open(CALENDAR_PATH, encoding='utf-8') as f:
            return json.load(f), None
    except (json.JSONDecodeError, IOError):
        return [], None


@app.route('/meetings/<path:filename>')
def meeting_detail(filename):
    """Show and optionally edit a meeting summary markdown file."""
    # Security: only allow filenames that look like YYYY-MM-DD-*.md
    if not re.match(r'^\d{4}-\d{2}-\d{2}-[\w\-]+\.md$', filename):
        abort(404)
    fp = os.path.join(MEETINGS_DIR, filename)
    if not os.path.exists(fp):
        abort(404)
    with open(fp, encoding='utf-8') as f:
        raw = f.read()

    # Parse header fields for the template
    meeting_date = ''
    source_url = ''
    source_label = ''
    attendees = ''
    title = ''

    for line in raw.splitlines():
        if line.startswith('# ') and not title:
            title = line[2:].strip()
        elif line.startswith('**Date:**'):
            meeting_date = line.split(':', 1)[1].strip()
        elif line.startswith('**Attendees:**'):
            attendees = line.split(':', 1)[1].strip()
        elif line.startswith('**Source:**'):
            um = re.search(r'\[([^\]]+)\]\((https?://[^\)]+)\)', line)
            if um:
                source_label = um.group(1)
                source_url = um.group(2)

    rendered_html = _render_meeting_markdown(raw)
    return render_template(
        'meeting_detail.html',
        filename=filename,
        title=title or filename,
        meeting_date=meeting_date,
        attendees=attendees,
        source_url=source_url,
        source_label=source_label,
        raw_content=raw,
        rendered_html=rendered_html,
    )


@app.route('/meetings/<path:filename>/save', methods=['POST'])
def meeting_save(filename):
    """Save updated meeting content."""
    if not re.match(r'^\d{4}-\d{2}-\d{2}-[\w\-]+\.md$', filename):
        abort(404)
    fp = os.path.join(MEETINGS_DIR, filename)
    if not os.path.exists(fp):
        abort(404)
    data = request.get_json(force=True)
    content = data.get('content', '')
    with open(fp, 'w', encoding='utf-8') as f:
        f.write(content)
    return jsonify({'ok': True})


@app.route('/api/calendar/refresh', methods=['POST'])
def api_calendar_refresh():
    """Fetch today's calendar events from Microsoft Graph and update dashboard_calendar.json."""
    try:
        from auth.graph_auth import get_access_token, _load_cache, _build_app, SCOPES
        from sources.ms_graph import get_today_events
    except ImportError as e:
        return jsonify({'ok': False, 'error': f'Import error: {e}'}), 500

    # Check if we have a valid cached token first (non-blocking)
    try:
        cache = _load_cache()
        app = _build_app(cache)
        accounts = app.get_accounts()

        if not accounts:
            return jsonify({
                'ok': False,
                'error': 'Not authenticated. Run `/productivity:update` from command line to authenticate with Microsoft Graph.',
                'needsAuth': True
            }), 401

        # Try silent token acquisition
        result = app.acquire_token_silent(SCOPES, account=accounts[0])
        if not result or "access_token" not in result:
            return jsonify({
                'ok': False,
                'error': 'Token expired. Run `/productivity:update` from command line to re-authenticate.',
                'needsAuth': True
            }), 401

        token = result["access_token"]
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Auth check failed: {e}'}), 500

    try:
        raw_events = get_today_events(token)
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Graph API error: {e}'}), 500

    # Convert Graph events to dashboard format
    try:
        from zoneinfo import ZoneInfo
        pacific = ZoneInfo("America/Los_Angeles")
    except Exception:
        pacific = None

    formatted = []
    for evt in raw_events:
        if evt.get('is_all_day'):
            continue

        start_raw = evt.get('start', '')
        end_raw = evt.get('end', '')
        tz_name = evt.get('timezone', 'UTC')

        def _parse_dt(s, tz_fallback):
            if not s:
                return None
            # Strip trailing Z and parse
            s = s.rstrip('Z')
            try:
                dt = datetime.fromisoformat(s)
            except ValueError:
                return None
            # Attach timezone
            if dt.tzinfo is None:
                try:
                    from zoneinfo import ZoneInfo
                    dt = dt.replace(tzinfo=ZoneInfo(tz_fallback))
                except Exception:
                    dt = dt.replace(tzinfo=timezone.utc)
            return dt

        start_dt = _parse_dt(start_raw, tz_name)
        end_dt = _parse_dt(end_raw, tz_name)

        if start_dt and pacific:
            start_dt = start_dt.astimezone(pacific)
        if end_dt and pacific:
            end_dt = end_dt.astimezone(pacific)

        def _fmt_time(dt):
            if not dt:
                return ''
            h = dt.hour % 12 or 12
            ampm = 'AM' if dt.hour < 12 else 'PM'
            return f"{h}:{dt.minute:02d} {ampm}"

        time_str = f"{_fmt_time(start_dt)} – {_fmt_time(end_dt)}" if start_dt and end_dt else ''

        attendees = ', '.join(
            a['name'] or a['email']
            for a in evt.get('attendees', [])
            if a.get('name') or a.get('email')
        )

        # Store end_time as ISO string for JavaScript comparison
        end_time_iso = end_dt.isoformat() if end_dt else ''

        # Strip (Past) and (Future) labels from title
        title = evt.get('subject', '')
        title = re.sub(r'\s*\((Past|Future)\)\s*$', '', title).strip()

        formatted.append({
            'time': time_str,
            'title': title,
            'attendees': attendees,
            'location': evt.get('location', ''),
            'end_time': end_time_iso,
        })

    # Write to disk
    try:
        with open(CALENDAR_PATH, 'w', encoding='utf-8') as f:
            json.dump(formatted, f, ensure_ascii=False)
    except IOError as e:
        return jsonify({'ok': False, 'error': f'Could not write calendar: {e}'}), 500

    return jsonify({'ok': True, 'events': formatted, 'count': len(formatted)})


@app.route('/api/task/complete', methods=['POST'])
def task_complete():
    data = request.get_json(force=True)
    task_text = data.get('text', '').strip()
    if not task_text or not os.path.exists(TASKS_PATH):
        return jsonify({'ok': False}), 400
    with open(TASKS_PATH, 'r', encoding='utf-8') as f:
        content = f.read()
    content = content.replace(f'- [ ] {task_text}', f'- [x] {task_text}', 1)
    with open(TASKS_PATH, 'w', encoding='utf-8') as f:
        f.write(content)
    return jsonify({'ok': True})


@app.route('/api/task/add', methods=['POST'])
def task_add():
    data = request.get_json(force=True)
    text = data.get('text', '').strip()
    priority = data.get('priority', 'Med').strip()
    section = data.get('section', 'Active').strip()
    if not text or not os.path.exists(TASKS_PATH):
        return jsonify({'ok': False}), 400
    with open(TASKS_PATH, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    new_line = f'- [ ] **[{priority}]** {text}\n'
    target = f'## {section}'
    inserted = False
    for i, ln in enumerate(lines):
        if ln.strip() == target:
            lines.insert(i + 1, new_line)
            inserted = True
            break
    if not inserted:
        lines.append(new_line)
    with open(TASKS_PATH, 'w', encoding='utf-8') as f:
        f.writelines(lines)
    return jsonify({'ok': True})


@app.route('/api/task/status', methods=['PATCH'])
def task_status_update():
    from sources.memory_reader import update_task_status
    data = request.get_json(force=True)
    section = data.get('section', '').strip()
    task_text = data.get('task_text', '').strip()
    new_status = data.get('new_status', '').strip()

    if not section or not task_text or not new_status:
        return jsonify({'success': False, 'error': 'section, task_text, and new_status are required'}), 400

    if new_status not in ['New', 'In Progress', 'Complete']:
        return jsonify({'success': False, 'error': 'Invalid status'}), 400

    success = update_task_status(section, task_text, new_status)

    if success:
        return jsonify({'success': True, 'new_status': new_status})
    else:
        return jsonify({'success': False, 'error': 'Task not found'}), 404


# ---------------------------------------------------------------------------
# CRM Blueprint
# ---------------------------------------------------------------------------

crm_bp = Blueprint('crm', __name__, url_prefix='/crm')

EDITABLE_FIELDS = {
    'stage', 'urgent', 'target', 'assigned_to', 'notes', 'closing'
}


# --- KB people ---

def parse_kb_person_file(path: str) -> dict:
    """Parse a memory/people/*.md file into a person dict."""
    slug = os.path.splitext(os.path.basename(path))[0]
    # Convert slug to Title Case as fallback name
    name_fallback = ' '.join(w.capitalize() for w in slug.replace('-', ' ').split())
    fields = {'name': name_fallback, 'org': '', 'title': '', 'email': '', 'phone': '', 'company': ''}
    field_map = {
        'name':         'name',
        'organization': 'company',
        'org':          'company',
        'company':      'company',
        'title':        'title',
        'role':         'title',
        'email':        'email',
        'phone':        'phone',
        'cell':         'phone',
        'mobile':       'phone',
    }
    try:
        with open(path, 'r', encoding='utf-8') as f:
            for line in f:
                # Parse # Heading as name
                h1 = re.match(r'^#\s+(.+)', line.strip())
                if h1:
                    fields['name'] = h1.group(1).strip()
                    continue
                # Parse **Field:** value format (with or without dash/bullet)
                m = re.match(r'(?:-|\*)?\s*\*\*([^:]+):\*\*\s*(.*)', line.strip())
                if m:
                    key = m.group(1).lower().strip()
                    val = m.group(2).strip()
                    if key in field_map and val:
                        fields[field_map[key]] = val
    except Exception:
        pass
    fields['slug'] = slug
    # Keep org field for backward compatibility (alias to company)
    fields['org'] = fields['company']
    return fields


@crm_bp.route('/api/kb-people')
def api_kb_people():
    q = request.args.get('q', '').lower().strip()
    config = load_crm_config()
    arec_team = {name.lower() for name in config.get('team', [])}
    people_dir = os.path.join(PROJECT_ROOT, 'memory', 'people')
    results = []
    if os.path.isdir(people_dir):
        for fname in sorted(os.listdir(people_dir)):
            if not fname.endswith('.md'):
                continue
            path = os.path.join(people_dir, fname)
            person = parse_kb_person_file(path)
            if person['name'].lower() in arec_team:
                continue
            if not q or q in person['name'].lower():
                results.append(person)
    results.sort(key=lambda p: p['name'].lower())
    return jsonify(results)


# --- Page routes ---

@crm_bp.route('/person/<slug>')
def person_detail(slug):
    """Person detail page — shows memory file + email history for a contact."""
    people_dir = os.path.join(PROJECT_ROOT, 'memory', 'people')
    path = os.path.join(people_dir, f'{slug}.md')
    if not os.path.exists(path):
        abort(404)
    person = parse_kb_person_file(path)
    with open(path, 'r', encoding='utf-8') as f:
        raw_content = f.read()
    emails = get_emails_for_org(person['org']) if person.get('org') else []
    # Filter to emails from/to this specific person
    if person.get('email'):
        em = person['email'].lower()
        person_emails = [
            e for e in emails
            if em in e.get('from', '').lower()
            or em in ' '.join(e.get('to', [])).lower()
            or em in ' '.join(e.get('cc', [])).lower()
        ]
    else:
        person_emails = []
    return render_template('crm_person_detail.html',
                           person=person,
                           slug=slug,
                           raw_content=raw_content,
                           person_emails=person_emails[:20])


@crm_bp.route('/')
@crm_bp.route('')
def pipeline():
    config = load_crm_config()
    offerings = load_offerings()
    return render_template('crm_pipeline.html', config=config, offerings=offerings)


@crm_bp.route('/people')
def people_list():
    return render_template('crm_people.html')


@crm_bp.route('/people/<slug>')
def people_person_detail(slug):
    """Person detail at /crm/people/<slug> — upgraded AI brief page."""
    return _render_person_detail(slug)


def _render_person_detail(slug):
    """Shared handler for person detail pages."""
    people_dir = os.path.join(PROJECT_ROOT, 'memory', 'people')
    path = os.path.join(people_dir, f'{slug}.md')
    if not os.path.exists(path):
        abort(404)
    person = parse_kb_person_file(path)
    return render_template('crm_person_detail.html', person=person, slug=slug)


@crm_bp.route('/orgs')
def orgs_list():
    return render_template('crm_orgs.html')


@crm_bp.route('/org/<path:name>/edit')
def org_edit(name):
    """Org editing page — contacts, type, notes, prospects."""
    config = load_crm_config()
    offerings = load_offerings()
    return render_template('crm_org_detail.html', org_name=name, config=config, offerings=offerings)


@crm_bp.route('/org/<path:name>')
def org_detail(name):
    """Redirect org detail to the prospect detail page for this org's primary/best prospect."""
    prospects = get_prospects_for_org(name)

    if not prospects:
        return redirect(url_for('crm.orgs_list'))

    if len(prospects) == 1:
        p = prospects[0]
        return redirect(url_for('crm.prospect_detail',
            offering=p.get('offering', ''), org=name))

    # Multiple prospects — pick highest stage, then most recent last_touch
    def _prospect_sort_key(p):
        stage = p.get('Stage', '0. Unknown')
        try:
            stage_num = int(stage.split('.')[0])
        except (ValueError, IndexError):
            stage_num = 0
        last_touch = p.get('Last Touch', '') or p.get('last_touch', '') or '2000-01-01'
        return (stage_num, last_touch)

    best = max(prospects, key=_prospect_sort_key)
    return redirect(url_for('crm.prospect_detail',
        offering=best.get('offering', ''), org=name))


@crm_bp.route('/prospect/<offering>/<path:org>')
def prospect_edit(offering, org):
    prospect = get_prospect(org, offering)
    if not prospect:
        abort(404)
    config = load_crm_config()
    contacts = get_contacts_for_org(org)

    # All known non-AREC people not already linked to this org
    arec_names = {member['name'].lower() for member in config['team']}
    org_slugs = {c['slug'] for c in contacts}
    other_contacts = [
        p for p in load_all_persons()
        if p['slug'] not in org_slugs
        and p['name'].lower() not in arec_names
        and 'arec' not in (p.get('organization') or '').lower()
        and 'avila real estate' not in (p.get('organization') or '').lower()
    ]

    return render_template('crm_prospect_edit.html',
                           prospect=prospect,
                           config=config,
                           contacts=contacts,
                           other_contacts=other_contacts,
                           offering=offering,
                           org=org)


@crm_bp.route('/prospect/<offering>/<path:org>/detail')
def prospect_detail(offering, org):
    """Prospect detail page — AI brief hero + reference sections."""
    prospect = get_prospect(org, offering)
    if not prospect:
        abort(404)
    config = load_crm_config()

    # Normalize urgency to a boolean for clean template rendering
    urgent_raw = prospect.get('Urgent', '') or prospect.get('urgent', '')
    prospect['urgent_bool'] = str(urgent_raw).strip().lower() in ('yes', 'true', 'high', '1')

    return render_template('crm_prospect_detail.html',
                           prospect=prospect,
                           config=config,
                           offering=offering,
                           org=org)


# --- Data API ---


@crm_bp.route('/api/prospect/<offering>/<path:org>/brief', methods=['GET', 'POST'])
def api_prospect_brief(offering, org):
    """Return structured relationship intelligence from all KB sources + content hash."""
    raw_data = collect_relationship_data(org, offering, base_dir=PROJECT_ROOT)
    content_hash = compute_content_hash(raw_data)
    return jsonify({**raw_data, 'content_hash': content_hash})


@crm_bp.route('/api/synthesize-brief', methods=['POST'])
def api_synthesize_brief():
    """Call Claude API to synthesize a narrative relationship brief from raw data."""
    data = request.get_json(force=True)
    org = data.get('org', '').strip()
    offering = data.get('offering', '').strip()

    if not org or not offering:
        return jsonify({'error': 'org and offering required'}), 400

    raw_data = collect_relationship_data(org, offering, base_dir=PROJECT_ROOT)
    content_hash = compute_content_hash(raw_data)
    context_block = build_context_block(raw_data)

    try:
        client = anthropic.Anthropic()
        message = client.messages.create(
            model='claude-sonnet-4-6',
            max_tokens=1500,
            system=BRIEF_SYSTEM_PROMPT,
            messages=[{
                'role': 'user',
                'content': (
                    f"Generate a relationship brief for {org} regarding {offering}.\n\n"
                    f"{context_block}"
                )
            }]
        )
        narrative = message.content[0].text
    except Exception:
        narrative = build_fallback_summary(raw_data)

    return jsonify({
        'narrative': narrative,
        'content_hash': content_hash,
    })


@crm_bp.route('/api/emails/<path:org>')
def api_emails_for_org(org):
    """Return paginated emails for an org from email_log.json."""
    limit = request.args.get('limit', 20, type=int)
    offset = request.args.get('offset', 0, type=int)
    emails = get_emails_for_org(org)
    paginated = emails[offset:offset + limit]
    return jsonify({
        'emails': paginated,
        'total': len(emails),
        'offset': offset,
        'limit': limit,
    })


@crm_bp.route('/api/email/<path:message_id>')
def api_email_detail(message_id):
    """Return a single email entry from the log."""
    email = find_email_by_message_id(message_id)
    if not email:
        abort(404)
    return jsonify(email)


@crm_bp.route('/api/prospect/<offering>/<path:org>/email-scan', methods=['POST'])
def api_prospect_email_scan(offering, org):
    """
    Deep email scan for a specific org — searches Archive + Sent Items over
    the last 90 days for any email related to the org's domain or contacts.
    Adds new matches to email_log.json (deduped). Returns count added.
    """
    from auth.graph_auth import get_access_token
    from sources.ms_graph import search_emails_deep
    from sources.crm_reader import get_org_domains, add_emails_to_log

    # Resolve domain for this org
    org_domains = get_org_domains()
    domain = ''
    for org_name, d in org_domains.items():
        if org_name.lower() == org.lower():
            domain = d
            break

    # Get known contact emails for the org
    contacts = get_contacts_for_org(org)
    contact_emails = [c.get('email', '') for c in contacts if c.get('email')]

    if not domain and not contact_emails:
        return jsonify({
            'error': (
                'No domain or contacts found for this org. '
                'Add a Domain field to the org in organizations.md first.'
            ),
            'added': 0,
        }), 400

    # Acquire MS Graph token (uses cached token — no device flow unless expired)
    try:
        token = get_access_token()
    except Exception as e:
        return jsonify({'error': f'MS Graph auth failed: {e}', 'added': 0}), 500

    # Run the deep search
    try:
        raw_emails = search_emails_deep(token, domain, contact_emails, days_back=90)
    except Exception as e:
        return jsonify({'error': f'Email search failed: {e}', 'added': 0}), 500

    if not raw_emails:
        return jsonify({
            'added': 0,
            'total': 0,
            'message': 'No matching emails found in the last 90 days.',
        })

    # Summarize each email via Claude Haiku and build log entries
    log_entries = []
    for email in raw_emails:
        subject = email.get('subject', '')
        preview = email.get('preview', '')
        from_addr = email.get('from', '')
        is_sent = email.get('isSent', False)

        # Generate a concise 1-2 sentence summary
        try:
            resp = client.messages.create(
                model='claude-haiku-4-5-20251001',
                max_tokens=100,
                messages=[{
                    'role': 'user',
                    'content': (
                        f"Summarize this {'outgoing' if is_sent else 'incoming'} "
                        f"real estate investor email in 1-2 sentences. "
                        f"Focus on the key action, commitment, or decision. Be specific.\n\n"
                        f"Subject: {subject}\n"
                        f"From: {from_addr}\n"
                        f"Preview: {preview}"
                    ),
                }],
            )
            summary = resp.content[0].text.strip()
        except Exception:
            summary = f"{'Sent' if is_sent else 'Received'}: {subject}"

        log_entries.append({
            'messageId': email.get('messageId'),
            'date': email.get('date', ''),
            'timestamp': email.get('timestamp', ''),
            'subject': subject,
            'from': from_addr,
            'fromName': email.get('fromName', ''),
            'to': email.get('to', []),
            'orgMatch': org,
            'matchType': 'deep-scan',
            'confidence': 0.90,
            'summary': summary,
            'outlookUrl': '',
        })

    added = add_emails_to_log(log_entries)
    already_logged = len(raw_emails) - added

    return jsonify({
        'added': added,
        'total': len(raw_emails),
        'message': (
            f'Scan complete — {added} new email{"s" if added != 1 else ""} added'
            + (f' ({already_logged} already logged).' if already_logged else '.')
        ),
    })


# --- Person API ---

@crm_bp.route('/api/person-data')
def api_person_data():
    """Collect all data about a person across all KB sources."""
    name = request.args.get('name', '').strip()
    if not name:
        return jsonify({'error': 'name required'}), 400

    data = collect_person_data(name, base_dir=PROJECT_ROOT)
    content_hash = hashlib.md5(
        json.dumps(data, sort_keys=True, default=str).encode()
    ).hexdigest()[:12]
    return jsonify({**data, 'content_hash': content_hash})


@crm_bp.route('/api/synthesize-person-brief', methods=['POST'])
def api_synthesize_person_brief():
    """Synthesize a person-focused AI narrative brief."""
    data = request.get_json(force=True)
    person_name = data.get('name', '').strip()
    if not person_name:
        return jsonify({'error': 'name required'}), 400

    raw_data = collect_person_data(person_name, base_dir=PROJECT_ROOT)
    content_hash = hashlib.md5(
        json.dumps(raw_data, sort_keys=True, default=str).encode()
    ).hexdigest()[:12]
    context_block = build_person_context_block(raw_data)

    try:
        client = anthropic.Anthropic()
        message = client.messages.create(
            model='claude-sonnet-4-6',
            max_tokens=1500,
            system=PERSON_BRIEF_SYSTEM_PROMPT,
            messages=[{
                'role': 'user',
                'content': (
                    f"Generate a person brief for {person_name}.\n\n"
                    f"{context_block}"
                )
            }]
        )
        narrative = message.content[0].text
    except Exception:
        narrative = build_person_fallback_summary(raw_data)

    return jsonify({
        'narrative': narrative,
        'content_hash': content_hash,
    })


@crm_bp.route('/api/person-update', methods=['POST'])
def api_person_update():
    """Accept free-text context about a person, AI routes updates to data stores."""
    data = request.get_json(force=True)
    person_name = data.get('name', '').strip()
    user_input = data.get('input', '').strip()
    if not person_name or not user_input:
        return jsonify({'error': 'name and input required'}), 400

    raw_data = collect_person_data(person_name, base_dir=PROJECT_ROOT)
    context_block = build_person_context_block(raw_data)
    config = load_crm_config()
    org_name = raw_data.get('org_name', '')

    try:
        client = anthropic.Anthropic()
        message = client.messages.create(
            model='claude-sonnet-4-6',
            max_tokens=2000,
            system=PERSON_UPDATE_ROUTING_PROMPT,
            messages=[{
                'role': 'user',
                'content': (
                    f"CURRENT PERSON DATA:\n{context_block}\n\n"
                    f"PERSON NAME: {person_name}\n"
                    f"ORGANIZATION: {org_name}\n"
                    f"VALID STAGES: {', '.join(config.get('stages', []))}\n"
                    f"TODAY'S DATE: {date.today().isoformat()}\n\n"
                    f"USER UPDATE:\n{user_input}\n\n"
                    "Determine what data store updates are needed and return JSON."
                )
            }]
        )
        response_text = message.content[0].text
        clean = response_text.replace('```json', '').replace('```', '').strip()
        updates = json.loads(clean)
    except Exception as e:
        return jsonify({'error': f'AI routing failed: {str(e)}'}), 500

    results = execute_person_updates(person_name, org_name, updates, base_dir=PROJECT_ROOT)

    updated_data = collect_person_data(person_name, base_dir=PROJECT_ROOT)
    new_hash = hashlib.md5(
        json.dumps(updated_data, sort_keys=True, default=str).encode()
    ).hexdigest()[:12]

    return jsonify({'actions': results, 'new_content_hash': new_hash})


@crm_bp.route('/people/api/<slug>/contact', methods=['PATCH'])
def api_person_contact_update(slug):
    """Update person contact fields (email, phone, title, company)."""
    data = request.get_json(force=True)

    people_dir = os.path.join(PROJECT_ROOT, 'memory', 'people')
    path = os.path.join(people_dir, f'{slug}.md')
    if not os.path.exists(path):
        return jsonify({'error': 'Person not found'}), 404

    # Read current file
    with open(path, 'r', encoding='utf-8') as f:
        content = f.read()

    lines = content.split('\n')

    # Track which fields we need to update/insert
    field_updates = {}
    for field in ('email', 'phone', 'title', 'company'):
        if field in data:
            field_updates[field] = data[field].strip()

    # Canonical field names for writing
    field_names = {
        'email': 'Email',
        'phone': 'Phone',
        'title': 'Title',
        'company': 'Company',
    }

    # Parse existing structured section (lines starting with **)
    structured_end = 0
    for i, line in enumerate(lines):
        if line.strip().startswith('**') and ':' in line:
            structured_end = i + 1
        elif line.strip().startswith('#'):
            if i > 0:  # Not the first heading
                break
        elif line.strip().startswith('##'):
            break

    # Update or insert fields in structured section
    updated_fields = set()
    for i in range(min(structured_end, len(lines))):
        line = lines[i]
        for field, value in field_updates.items():
            pattern = rf'^\*\*({field}|{field_names[field]}|{"Organization" if field == "company" else ""}|{"Role" if field == "title" else ""}|{"Cell" if field == "phone" else ""}|{"Mobile" if field == "phone" else ""}):\*\*'
            if re.match(pattern, line.strip(), re.IGNORECASE):
                if value:
                    lines[i] = f'**{field_names[field]}:** {value}'
                else:
                    lines[i] = ''  # Remove if blank
                updated_fields.add(field)

    # Insert any fields that weren't found
    insert_point = structured_end if structured_end > 0 else 2  # After # Name line
    for field in ('company', 'title', 'email', 'phone'):  # Order matters
        if field in field_updates and field not in updated_fields:
            value = field_updates[field]
            if value:  # Only insert non-empty fields
                lines.insert(insert_point, f'**{field_names[field]}:** {value}')
                insert_point += 1

    # Clean up empty lines
    lines = [ln for ln in lines if ln or ln == '\n']

    # Write back
    with open(path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

    # Return updated person data
    person = parse_kb_person_file(path)
    return jsonify(person)

@crm_bp.route('/api/offerings')
def api_offerings():
    return jsonify(load_offerings())


@crm_bp.route('/api/prospects')
def api_prospects():
    offering = request.args.get('offering', '')
    include_closed = request.args.get('include_closed', 'false').lower() == 'true'
    prospects = load_prospects(offering if offering else None)
    if not include_closed:
        excluded = {'8. Closed', '0. Not Pursuing', '0. Declined'}
        prospects = [p for p in prospects if p.get('Stage', '') not in excluded]

    # Enrich with tasks from TASKS.md (old (OrgName) format — backward compat)
    tasks_by_org = load_tasks_by_org()
    # New [org: ...] tag format tasks — keyed by org name (case-sensitive from file)
    all_new_tasks = get_all_prospect_tasks()
    new_tasks_by_org: dict = {}
    for t in all_new_tasks:
        new_tasks_by_org.setdefault(t['org'], []).append(
            {k: v for k, v in t.items() if k != 'org'}
        )

    for p in prospects:
        org_name = p.get('org', '')
        org_tasks = tasks_by_org.get(org_name, [])
        # Flat string for display/backward compat
        if org_tasks:
            p['Tasks'] = ' | '.join(
                f"[@{t['owner']}] {t['task']}" for t in org_tasks
            )
        else:
            p['Tasks'] = ''
        # Structured data for modal editing (old format)
        p['_tasks'] = org_tasks
        # New-format prospect tasks
        new_tasks = new_tasks_by_org.get(org_name, [])
        p['prospect_tasks'] = new_tasks
        p['open_task_count'] = sum(1 for t in new_tasks if t['status'] == 'open')

    return jsonify(prospects)


@crm_bp.route('/api/fund-summary')
def api_fund_summary():
    offering = request.args.get('offering', '')
    if offering:
        return jsonify(get_fund_summary(offering))
    return jsonify(get_fund_summary_all())


# --- Inline field edit ---

@crm_bp.route('/api/prospect/field', methods=['PATCH'])
def api_patch_prospect_field():
    data = request.get_json(force=True)
    org = data.get('org', '').strip()
    offering = data.get('offering', '').strip()
    field = data.get('field', '').strip().lower()
    raw_value = data.get('value', '')
    value = str(raw_value).strip() if isinstance(raw_value, str) else raw_value

    if not org or not offering or not field:
        return jsonify({'error': 'org, offering, and field are required'}), 400

    if field == 'next_action':
        return jsonify({'error': 'next_action field has been removed from the data model'}), 400

    if field not in EDITABLE_FIELDS:
        return jsonify({'error': f'Field "{field}" is not editable'}), 400

    config = load_crm_config()
    if field == 'stage' and value not in config['stages'] and value != '':
        return jsonify({'error': f'Invalid stage: {value}'}), 400
    if field == 'assigned_to' and value != '':
        valid_names = {member['name'] for member in config['team']}
        valid_names.update(m['short'] for m in config.get('team_map', []))
        valid_names.update(m['full'] for m in config.get('team_map', []))
        if value not in valid_names:
            return jsonify({'error': f'Invalid team member: {value}'}), 400
    if field == 'closing' and value not in config['closing_options'] and value != '':
        return jsonify({'error': f'Invalid closing option: {value}'}), 400

    if field == 'target':
        # Normalize to display format
        parsed = _parse_currency(value)
        value = f"${parsed:,.0f}" if parsed else '$0'

    update_prospect_field(org, offering, field, value)
    updated = get_prospect(org, offering)
    return jsonify(updated)


# --- Prospect Task API (TASKS.md [org: ...] format) ---

@crm_bp.route('/api/tasks', methods=['GET'])
def api_crm_tasks_list():
    org = request.args.get('org', '').strip()
    if not org:
        return jsonify({'error': 'org parameter required'}), 400
    return jsonify(get_tasks_for_prospect(org))


@crm_bp.route('/api/tasks', methods=['POST'])
def api_crm_tasks_create():
    data = request.get_json(force=True)
    org = data.get('org', '').strip()
    text = data.get('text', '').strip()
    owner = data.get('owner', '').strip()
    priority = data.get('priority', 'Med').strip()
    section = data.get('section', 'IR / Fundraising').strip()
    if not org or not text or not owner:
        return jsonify({'error': 'org, text, and owner are required'}), 400
    success = add_prospect_task(org, text, owner, priority, section)
    if not success:
        return jsonify({'error': 'Failed to add task'}), 500
    return jsonify({'ok': True}), 201


@crm_bp.route('/api/tasks/complete', methods=['PATCH'])
def api_crm_tasks_complete():
    data = request.get_json(force=True)
    org = data.get('org', '').strip()
    task_text = data.get('task_text', '').strip()
    if not org or not task_text:
        return jsonify({'error': 'org and task_text are required'}), 400
    success = complete_prospect_task(org, task_text)
    if not success:
        return jsonify({'error': 'Task not found'}), 404
    return jsonify({'ok': True})


# --- Org API ---

@crm_bp.route('/api/org/<path:name>', methods=['GET'])
def api_org_get(name):
    org = get_organization(name)
    if not org:
        # Return empty org structure rather than 404 (org may exist in prospects but not orgs)
        org = {'name': name, 'Type': '', 'Notes': ''}
    contacts = get_contacts_for_org(name)
    prospects = get_prospects_for_org(name)
    return jsonify({
        'org': org,
        'contacts': contacts,
        'prospects': prospects,
    })


@crm_bp.route('/api/org/<path:name>', methods=['PATCH'])
def api_org_patch(name):
    data = request.get_json(force=True)
    allowed = {'type', 'notes', 'domain'}
    payload = {}
    if 'type' in data:
        payload['Type'] = data['type']
    if 'domain' in data:
        payload['Domain'] = data['domain']
    if 'notes' in data:
        payload['Notes'] = data['notes']
    if not payload:
        return jsonify({'error': 'No valid fields to update'}), 400

    # Ensure org exists in organizations.md first
    existing = get_organization(name)
    if existing:
        merged = {**existing, **payload}
    else:
        merged = {'name': name, 'Type': '', 'Notes': '', **payload}

    write_organization(name, merged)
    return jsonify(get_organization(name) or merged)


# --- Contact API ---

@crm_bp.route('/api/contact', methods=['POST'])
def api_contact_create():
    data = request.get_json(force=True)
    name = data.get('name', '').strip()
    org = data.get('org', '').strip()
    if not name or not org:
        return jsonify({'error': 'name and org are required'}), 400

    email = data.get('email', '').strip()
    role = data.get('role', '').strip()
    person_type = data.get('type', 'investor').strip()

    slug = create_person_file(name, org, email, role, person_type)
    person = load_person(slug)
    return jsonify(person), 201


@crm_bp.route('/api/contact/<path:org_and_name>', methods=['PATCH'])
def api_contact_patch(org_and_name):
    # URL format: /api/contact/<org>/<name>
    # Split on last slash
    parts = org_and_name.rsplit('/', 1)
    if len(parts) != 2:
        return jsonify({'error': 'URL must be /api/contact/<org>/<name>'}), 400
    org, name = parts[0], parts[1]
    data = request.get_json(force=True)
    allowed_fields = {'role', 'email', 'phone', 'title'}
    payload = {k: v for k, v in data.items() if k in allowed_fields}
    if not payload:
        return jsonify({'error': 'No valid fields to update'}), 400
    success = update_contact_fields(org, name, payload)
    if not success:
        return jsonify({'error': 'Contact not found'}), 404
    return jsonify({'ok': True})


# --- Prospect full save (edit page) ---

@crm_bp.route('/api/prospect/save', methods=['POST'])
def api_prospect_save():
    data = request.get_json(force=True)
    org = data.get('org', '').strip()
    offering = data.get('offering', '').strip()
    fields = data.get('fields', {})
    if not org or not offering:
        return jsonify({'error': 'org and offering are required'}), 400
    if not get_prospect(org, offering):
        return jsonify({'error': 'Prospect not found'}), 404
    for field, value in fields.items():
        update_prospect_field(org, offering, field, str(value))
    # Ensure last_touch is set to today
    update_prospect_field(org, offering, 'last_touch', date.today().isoformat())
    return jsonify({'status': 'ok'})


# --- Prospect create ---

@crm_bp.route('/api/prospect', methods=['POST'])
def api_prospect_create():
    data = request.get_json(force=True)
    org = data.get('org', '').strip()
    offering = data.get('offering', '').strip()
    if not org or not offering:
        return jsonify({'error': 'org and offering are required'}), 400

    # Don't create duplicate
    existing = get_prospect(org, offering)
    if existing:
        return jsonify({'error': 'Prospect already exists for this org + offering'}), 409

    new_prospect = {
        'org': org,
        'offering': offering,
        'Stage': data.get('stage', '1. Prospect'),
        'Target': data.get('target', '$0'),
        'Committed': '$0',
        'Primary Contact': '',
        'Closing': '',
        'Urgent': '',
        'Assigned To': '',
        'Notes': '',
        'Last Touch': '',
    }
    write_prospect(org, offering, new_prospect)
    created = get_prospect(org, offering)
    return jsonify(created), 201


@crm_bp.route('/api/prospect', methods=['DELETE'])
def api_prospect_delete():
    data = request.get_json(force=True)
    org = data.get('org', '').strip()
    offering = data.get('offering', '').strip()
    if not org or not offering:
        return jsonify({'error': 'org and offering required'}), 400
    try:
        delete_prospect(org, offering)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# --- Unmatched review ---

@crm_bp.route('/api/unmatched', methods=['GET'])
def api_unmatched_list():
    return jsonify(load_unmatched())


@crm_bp.route('/api/unmatched/resolve', methods=['POST'])
def api_unmatched_resolve():
    """Resolve an unmatched item by logging it as an interaction."""
    data = request.get_json(force=True)
    email = data.get('participant_email', '').strip()
    org = data.get('org', '').strip()
    offering = data.get('offering', '').strip()
    int_type = data.get('type', 'Email').strip()
    subject = data.get('subject', '').strip()
    int_date = data.get('date', date.today().isoformat())

    if not email or not org:
        return jsonify({'error': 'participant_email and org required'}), 400

    # Log the interaction
    append_interaction({
        'org': org,
        'type': int_type,
        'offering': offering,
        'date': int_date,
        'contact': data.get('participant_name', ''),
        'subject': subject,
        'summary': f'Manual resolve: {subject}',
        'source': 'manual',
    })

    # Remove from unmatched queue
    remove_unmatched(email)

    return jsonify({'ok': True})


@crm_bp.route('/api/unmatched/<path:email>', methods=['DELETE'])
def api_unmatched_dismiss(email):
    remove_unmatched(email)
    return jsonify({'ok': True})


@crm_bp.route('/api/auto-capture', methods=['POST'])
def api_auto_capture():
    """Manual trigger for Graph auto-capture."""
    try:
        from auth.graph_auth import get_access_token
        from sources.crm_graph_sync import run_auto_capture

        token = get_access_token()
        stats = run_auto_capture(token)
        return jsonify({
            'ok': True,
            'emails_scanned': stats.get('matched', 0) + stats.get('unmatched', 0) + stats.get('skipped_dedup', 0),
            'meetings_scanned': 0,  # Not tracked separately in current impl
            'interactions_logged': stats.get('matched', 0),
            'prospects_touched': stats.get('matched', 0),
            'duplicates_skipped': stats.get('skipped_dedup', 0),
            'unmatched_count': stats.get('unmatched', 0),
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# --- Org list (for dropdowns) ---

@crm_bp.route('/api/orgs')
def api_orgs():
    orgs = load_organizations()
    return jsonify([o['name'] for o in orgs])


@crm_bp.route('/api/export')
def api_export_pipeline():
    """Export pipeline to Excel with formatting."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    offering = request.args.get('offering')
    if not offering:
        return jsonify({"error": "offering required"}), 400

    # Load prospects and apply filters
    prospects = load_prospects(offering)
    orgs = {o['name']: o for o in load_organizations()}

    # Apply filters from query params
    stage_filter = request.args.get('stage')
    urgent_only = request.args.get('urgent') == 'true'
    type_filter = request.args.get('type')
    assigned_filter = request.args.get('assigned')

    if stage_filter:
        prospects = [p for p in prospects if p.get('Stage') == stage_filter]
    if urgent_only:
        prospects = [p for p in prospects if p.get('Urgent')]
    if type_filter:
        prospects = [p for p in prospects if orgs.get(p.get('org', ''), {}).get('Type') == type_filter]
    if assigned_filter:
        prospects = [p for p in prospects if assigned_filter in str(p.get('Assigned To', ''))]

    # Sort: stage desc, urgency, target desc
    def sort_key(p):
        stage = p.get('Stage', '0. Unknown')
        stage_num = 0
        if stage and stage[0].isdigit():
            try:
                stage_num = int(stage.split('.')[0])
            except (ValueError, IndexError):
                stage_num = 0

        urgent = str(p.get('Urgent', '')).lower()
        urgency_order = {'yes': 0, 'high': 0, 'med': 1, 'medium': 1, 'low': 2, '': 3}
        urgency_val = urgency_order.get(urgent, 3)

        target = p.get('Target', '$0')
        target_val = _parse_currency(target) if target else 0

        return (-stage_num, urgency_val, -target_val)

    prospects.sort(key=sort_key)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Pipeline"

    # Define columns
    columns = [
        ('Organization', 30),
        ('Type', 16),
        ('Stage', 20),
        ('Target', 16),
        ('Committed', 16),
        ('Closing', 10),
        ('Urgency', 10),
        ('Assigned To', 20),
        ('Primary Contact', 22),
        ('Next Action', 35),
        ('Notes', 40),
        ('Last Touch', 14),
    ]

    # Set column widths
    for idx, (col_name, width) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Header row
    header_fill = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
    header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center')

    for idx, (col_name, _) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    ws.row_dimensions[1].height = 30

    # Enable auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Data rows
    row_fill_alt = PatternFill(start_color='f8f9fa', end_color='f8f9fa', fill_type='solid')
    data_font = Font(name='Arial', size=10)

    today = date.today()

    for row_idx, p in enumerate(prospects, start=2):
        org_name = p.get('org', '')
        org_data = orgs.get(org_name, {})
        org_type = org_data.get('Type', '')

        # Alternating row fill
        if row_idx % 2 == 0:
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = row_fill_alt

        # A: Organization (bold)
        cell_org = ws.cell(row=row_idx, column=1, value=org_name)
        cell_org.font = Font(name='Arial', size=10, bold=True)
        cell_org.alignment = Alignment(horizontal='left', vertical='top')

        # B: Type
        cell_type = ws.cell(row=row_idx, column=2, value=org_type)
        cell_type.font = data_font
        cell_type.alignment = Alignment(horizontal='left', vertical='top')

        # C: Stage
        stage = p.get('Stage', '')
        cell_stage = ws.cell(row=row_idx, column=3, value=stage)
        cell_stage.font = data_font
        cell_stage.alignment = Alignment(horizontal='left', vertical='top')

        # D: Target (currency, numeric)
        target_str = p.get('Target', '$0')
        target_val = _parse_currency(target_str) if target_str else 0
        cell_target = ws.cell(row=row_idx, column=4, value=target_val)
        cell_target.font = data_font
        cell_target.number_format = '$#,##0'
        cell_target.alignment = Alignment(horizontal='right', vertical='top')

        # E: Committed (currency, numeric)
        committed_str = p.get('Committed', '$0')
        committed_val = _parse_currency(committed_str) if committed_str else 0
        cell_committed = ws.cell(row=row_idx, column=5, value=committed_val)
        cell_committed.font = data_font
        cell_committed.number_format = '$#,##0'
        cell_committed.alignment = Alignment(horizontal='right', vertical='top')

        # F: Closing
        closing = p.get('Closing', '')
        cell_closing = ws.cell(row=row_idx, column=6, value=closing)
        cell_closing.font = data_font
        cell_closing.alignment = Alignment(horizontal='center', vertical='top')

        # G: Urgency (with conditional fill)
        urgency = p.get('Urgency', '').strip()
        cell_urgency = ws.cell(row=row_idx, column=7, value=urgency)
        cell_urgency.font = data_font
        cell_urgency.alignment = Alignment(horizontal='center', vertical='top')

        # Apply conditional formatting based on urgency level
        if urgency == 'High':
            cell_urgency.fill = PatternFill(start_color='fef2f2', end_color='fef2f2', fill_type='solid')
            cell_urgency.font = Font(name='Arial', size=10, color='ef4444')
        elif urgency == 'Med':
            cell_urgency.fill = PatternFill(start_color='fffbeb', end_color='fffbeb', fill_type='solid')
            cell_urgency.font = Font(name='Arial', size=10, color='f59e0b')
        elif urgency == 'Low':
            cell_urgency.fill = PatternFill(start_color='f9fafb', end_color='f9fafb', fill_type='solid')
            cell_urgency.font = Font(name='Arial', size=10, color='9ca3af')

        # H: Assigned To
        assigned = p.get('Assigned To', '')
        cell_assigned = ws.cell(row=row_idx, column=8, value=assigned)
        cell_assigned.font = data_font
        cell_assigned.alignment = Alignment(horizontal='left', vertical='top')

        # I: Primary Contact
        primary_contact = p.get('Primary Contact', '')
        cell_contact = ws.cell(row=row_idx, column=9, value=primary_contact)
        cell_contact.font = data_font
        cell_contact.alignment = Alignment(horizontal='left', vertical='top')

        # J: Next Action (wrap text)
        next_action = p.get('Next Action', '')
        cell_next = ws.cell(row=row_idx, column=10, value=next_action)
        cell_next.font = data_font
        cell_next.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        # K: Notes (wrap text)
        notes = p.get('Notes', '')
        cell_notes = ws.cell(row=row_idx, column=11, value=notes)
        cell_notes.font = data_font
        cell_notes.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        # L: Last Touch (with staleness fill)
        last_touch = p.get('Last Touch', '')
        cell_last_touch = ws.cell(row=row_idx, column=12, value=last_touch)
        cell_last_touch.font = data_font
        cell_last_touch.alignment = Alignment(horizontal='center', vertical='top')

        if last_touch:
            try:
                touch_date = datetime.strptime(last_touch, '%Y-%m-%d').date()
                days_ago = (today - touch_date).days

                if days_ago < 7:
                    cell_last_touch.fill = PatternFill(start_color='f0fdf4', end_color='f0fdf4', fill_type='solid')
                elif days_ago <= 14:
                    cell_last_touch.fill = PatternFill(start_color='fffbeb', end_color='fffbeb', fill_type='solid')
                else:
                    cell_last_touch.fill = PatternFill(start_color='fef2f2', end_color='fef2f2', fill_type='solid')
            except (ValueError, AttributeError):
                pass

    # Summary row
    summary_row = len(prospects) + 2
    cell_total_label = ws.cell(row=summary_row, column=1, value=f"TOTAL ({len(prospects)} prospects)")
    cell_total_label.font = Font(name='Arial', size=10, bold=True)

    # Target sum
    cell_target_sum = ws.cell(row=summary_row, column=4, value=f"=SUM(D2:D{summary_row-1})")
    cell_target_sum.font = Font(name='Arial', size=10, bold=True)
    cell_target_sum.number_format = '$#,##0'
    cell_target_sum.alignment = Alignment(horizontal='right')

    # Committed sum
    cell_committed_sum = ws.cell(row=summary_row, column=5, value=f"=SUM(E2:E{summary_row-1})")
    cell_committed_sum.font = Font(name='Arial', size=10, bold=True)
    cell_committed_sum.number_format = '$#,##0'
    cell_committed_sum.alignment = Alignment(horizontal='right')

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Generate filename
    safe_offering = offering.replace(' ', '_').replace('/', '_')
    filename = f"AREC_Pipeline_{safe_offering}_{today.isoformat()}.xlsx"

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@crm_bp.route('/api/org', methods=['POST'])
def api_org_create():
    """Create a new organization."""
    data = request.get_json(force=True)
    name = data.get('name', '').strip()
    org_type = data.get('type', '').strip()
    domain = data.get('domain', '').strip()
    notes = data.get('notes', '').strip()

    if not name:
        return jsonify({'error': 'Organization name is required'}), 400

    # Check if org already exists
    existing = get_organization(name)
    if existing:
        return jsonify({'error': 'Organization already exists'}), 409

    # Create the org
    org_data = {
        'name': name,
        'Type': org_type,
        'Domain': domain,
        'Notes': notes
    }
    write_organization(name, org_data)

    return jsonify({'ok': True, 'org': org_data}), 201


# --- Meeting History ---

@crm_bp.route('/api/org/<path:name>/meetings', methods=['GET'])
def api_org_meetings(name):
    meetings = load_meeting_history(name)
    return jsonify(meetings)


@crm_bp.route('/api/org/<path:name>/meetings', methods=['POST'])
def api_org_meeting_add(name):
    data = request.get_json(force=True)
    add_meeting_entry(
        org=name,
        date=data.get('date', ''),
        title=data.get('title', ''),
        attendees=data.get('attendees', ''),
        source=data.get('source', 'manual'),
        notion_url=data.get('notion_url', ''),
    )
    return jsonify({'ok': True})


app.register_blueprint(crm_bp)


# ---------------------------------------------------------------------------
# Tasks Blueprint
# ---------------------------------------------------------------------------

tasks_bp = Blueprint('tasks', __name__, url_prefix='/tasks')

TASK_SECTIONS = ['Fundraising - Me', 'Fundraising - Others', 'Other Work', 'Personal']


def _parse_task_line(line: str, section: str) -> dict:
    """Parse a single task markdown line into a dict."""
    line = line.rstrip()
    done = line.startswith('- [x] ')
    raw = line[6:].strip()
    text = raw

    # Priority
    priority = 'Med'
    pm = re.match(r'\*\*\[(\w+)\]\*\*\s*', text)
    if pm:
        priority = pm.group(1)
        text = text[pm.end():]

    # Status — based on checkbox and **[→]** marker
    if done:
        status = 'Complete'
    elif text.startswith('**[→]**'):
        status = 'In Progress'
        text = text[7:].strip()  # Remove **[→]** marker
    else:
        status = 'New'

    # Clean up old [STATUS:xxx] format for backward compatibility
    text = re.sub(r'\s*\[STATUS:\w+\]\s*', ' ', text).strip()

    # Completion date
    completion_date = None
    cdm = re.search(r'\s*—\s*completed\s+(\d{4}-\d{2}-\d{2})', text)
    if cdm:
        completion_date = cdm.group(1)
        text = text[:cdm.start()]

    # New: assigned:Name inline field — extract before splitting context
    assigned_to = None
    am = re.search(r'\s*—\s*assigned:([^—\n]+)', text)
    if am:
        assigned_to = am.group(1).strip()
        text = text[:am.start()] + text[am.end():]
        text = text.strip()

    # Fallback: legacy **@Name** format (for any pre-migration tasks)
    if assigned_to is None:
        om = re.match(r'\*\*@([^*]+)\*\*\s*', text)
        if om:
            assigned_to = om.group(1)
            text = text[om.end():]

    # Extract (OrgName) suffix from end of full text, before context split.
    # Filters out non-org parens like ($3M target) or (next week).
    org = ''
    org_m = re.search(r'\(([^)]+)\)\s*$', text)
    if org_m:
        candidate = org_m.group(1).strip()
        if not re.match(r'^[\$\d]', candidate):
            org = candidate
            text = text[:org_m.start()].rstrip()

    # Context (after —)
    context = ''
    di = text.find(' — ')
    if di >= 0:
        context = text[di + 3:]
        text = text[:di]

    # Strip ~~strikethrough~~
    text = re.sub(r'~~(.+?)~~', r'\1', text)

    return {
        'text': text.strip(),
        'priority': priority,
        'status': status,
        'context': context,
        'assigned_to': assigned_to,
        'org': org,
        'complete': done,
        'completion_date': completion_date,
        'raw': raw,
    }


def _load_tasks_full() -> dict:
    """
    Returns {section_name: [task_dicts_with_index]}.
    Sections: Fundraising - Me, Fundraising - Others, Other Work, Personal (+ Done).
    'index' = 0-based position of that task within the section.
    """
    if not os.path.exists(TASKS_PATH):
        return {s: [] for s in TASK_SECTIONS + ['Done']}

    result = {}
    current_name = None
    current_tasks = []

    def flush():
        if current_name is not None:
            result[current_name] = current_tasks[:]

    with open(TASKS_PATH, encoding='utf-8') as f:
        for line in f:
            m = re.match(r'^## (.+)$', line.rstrip())
            if m:
                flush()
                current_name = m.group(1).strip()
                current_tasks = []
                continue
            if current_name and (line.startswith('- [ ] ') or line.startswith('- [x] ')):
                task = _parse_task_line(line, current_name)
                task['index'] = len(current_tasks)
                task['section'] = current_name
                current_tasks.append(task)
    flush()

    # Ensure all canonical sections are present
    for s in TASK_SECTIONS + ['Done']:
        if s not in result:
            result[s] = []
    return result


def _read_task_lines() -> list:
    if not os.path.exists(TASKS_PATH):
        return []
    with open(TASKS_PATH, 'r', encoding='utf-8') as f:
        return f.readlines()


def _write_task_file(lines: list) -> None:
    with open(TASKS_PATH, 'w', encoding='utf-8') as f:
        f.writelines(lines)


def _find_task_line(lines: list, section: str, index: int):
    """
    Return the file line index of the task at position [index] within [section].
    Returns -1 if not found.
    """
    in_section = False
    count = 0
    for i, ln in enumerate(lines):
        m = re.match(r'^## (.+)$', ln.rstrip())
        if m:
            in_section = m.group(1).strip() == section
            continue
        if in_section and (ln.startswith('- [ ] ') or ln.startswith('- [x] ')):
            if count == index:
                return i
            count += 1
    return -1


def _format_task_line(text: str, priority: str, context: str,
                      assigned_to: str, section: str, done: bool = False,
                      completion_date: str = None, status: str = 'New', org: str = '') -> str:
    """Serialize a task dict back to a markdown line."""
    checkbox = '- [x] ' if done else '- [ ] '
    line = f'**[{priority}]** '

    # Add **[→]** for In Progress status (after priority, before text)
    if status == 'In Progress' and not done:
        line += '**[→]** '

    line += text
    # Add org tag if present
    if org:
        line += f' ({org})'
    if context:
        line += f' — {context}'
    if assigned_to:
        line += f' — assigned:{assigned_to}'
    if done and completion_date:
        line += f' — completed {completion_date}'
    return checkbox + line + '\n'


# --- Routes ---

@tasks_bp.route('/', methods=['GET'])
@tasks_bp.route('', methods=['GET'])
def tasks_page():
    config = load_crm_config()
    return render_template('tasks/tasks.html', config=config)


@tasks_bp.route('/api/tasks', methods=['GET'])
def api_tasks():
    return jsonify(_load_tasks_full())


@tasks_bp.route('/api/task', methods=['POST'])
def api_task_create():
    data = request.get_json(force=True)
    section = data.get('section', '').strip()
    text = data.get('text', '').strip()
    priority = data.get('priority', 'Med').strip()
    context = data.get('context', '').strip()
    assigned_to = data.get('assigned_to', '').strip()
    status = data.get('status', 'new').strip()
    org = data.get('org', '').strip()

    if not text or section not in TASK_SECTIONS:
        return jsonify({'ok': False, 'error': 'text and valid section required'}), 400

    new_line = _format_task_line(text, priority, context, assigned_to, section, status=status, org=org)

    lines = _read_task_lines()
    target = f'## {section}'
    inserted = False
    for i, ln in enumerate(lines):
        if ln.strip() == target:
            lines.insert(i + 1, new_line)
            inserted = True
            break
    if not inserted:
        lines.append(f'\n## {section}\n')
        lines.append(new_line)
    _write_task_file(lines)
    return jsonify({'ok': True})


@tasks_bp.route('/api/task/<section>/<int:index>', methods=['PUT'])
def api_task_update(section, index):
    data = request.get_json(force=True)
    new_section = data.get('section', section).strip()
    text = data.get('text', '').strip()
    priority = data.get('priority', 'Med').strip()
    context = data.get('context', '').strip()
    assigned_to = data.get('assigned_to', '').strip()
    status = data.get('status', 'open').strip()
    org = data.get('org', '').strip()

    if not text:
        return jsonify({'ok': False, 'error': 'text required'}), 400

    lines = _read_task_lines()
    li = _find_task_line(lines, section, index)
    if li == -1:
        return jsonify({'ok': False, 'error': 'task not found'}), 404

    original = lines[li]
    was_done = original.startswith('- [x] ')
    cdm = re.search(r'—\s*completed\s+(\d{4}-\d{2}-\d{2})', original)
    completion_date = cdm.group(1) if cdm else None

    new_line = _format_task_line(text, priority, context, assigned_to,
                                  new_section, done=was_done,
                                  completion_date=completion_date, status=status, org=org)

    if new_section == section:
        lines[li] = new_line
        _write_task_file(lines)
    else:
        # Move: remove from current section, insert into new section
        lines.pop(li)
        target = f'## {new_section}'
        inserted = False
        for i, ln in enumerate(lines):
            if ln.strip() == target:
                lines.insert(i + 1, new_line)
                inserted = True
                break
        if not inserted:
            lines.append(f'\n## {new_section}\n')
            lines.append(new_line)
        _write_task_file(lines)

    return jsonify({'ok': True})


@tasks_bp.route('/api/task/<section>/<int:index>', methods=['DELETE'])
def api_task_delete(section, index):
    lines = _read_task_lines()
    li = _find_task_line(lines, section, index)
    if li == -1:
        return jsonify({'ok': False, 'error': 'task not found'}), 404
    lines.pop(li)
    _write_task_file(lines)
    return jsonify({'ok': True})


@tasks_bp.route('/api/task/<section>/<int:index>/complete', methods=['POST'])
def api_task_complete_new(section, index):
    today = date.today().isoformat()
    lines = _read_task_lines()
    li = _find_task_line(lines, section, index)
    if li == -1:
        return jsonify({'ok': False, 'error': 'task not found'}), 404
    ln = lines[li]
    if not ln.startswith('- [ ] '):
        return jsonify({'ok': False, 'error': 'task already complete'}), 400
    ln = '- [x] ' + ln[6:]
    # Append completion date if not already there
    ln = ln.rstrip()
    if 'completed' not in ln:
        ln += f' — completed {today}'
    lines[li] = ln + '\n'
    _write_task_file(lines)
    return jsonify({'ok': True})


@tasks_bp.route('/api/task/<section>/<int:index>/restore', methods=['POST'])
def api_task_restore(section, index):
    lines = _read_task_lines()
    li = _find_task_line(lines, section, index)
    if li == -1:
        return jsonify({'ok': False, 'error': 'task not found'}), 404
    ln = lines[li]
    if not ln.startswith('- [x] '):
        return jsonify({'ok': False, 'error': 'task not complete'}), 400
    ln = '- [ ] ' + ln[6:]
    # Strip completion date
    ln = re.sub(r'\s*—\s*completed\s+\d{4}-\d{2}-\d{2}', '', ln.rstrip())
    lines[li] = ln + '\n'
    _write_task_file(lines)
    return jsonify({'ok': True})


@tasks_bp.route('/api/task/<section>/<int:index>/status', methods=['PATCH'])
def api_task_status_update(section, index):
    """Update task status inline (for status chip clicks)."""
    data = request.get_json(force=True)
    new_status = data.get('status', '').strip()

    valid_statuses = ['New', 'In Progress', 'Complete']
    if new_status not in valid_statuses:
        return jsonify({'ok': False, 'error': f'status must be one of: {", ".join(valid_statuses)}'}), 400

    lines = _read_task_lines()
    li = _find_task_line(lines, section, index)
    if li == -1:
        return jsonify({'ok': False, 'error': 'task not found'}), 404

    # Parse current task to preserve all fields
    task = _parse_task_line(lines[li], section)

    # Update status field, handle complete checkbox
    done = new_status == 'Complete'
    completion_date = date.today().isoformat() if done and not task['complete'] else task.get('completion_date')

    # Re-serialize with new status
    new_line = _format_task_line(
        task['text'],
        task['priority'],
        task['context'],
        task['assigned_to'],
        section,
        done=done,
        completion_date=completion_date,
        status=new_status,
        org=task.get('org', '')
    )

    lines[li] = new_line
    _write_task_file(lines)

    return jsonify({'ok': True, 'status': new_status})


@tasks_bp.route('/api/task/<section>/<int:index>/priority', methods=['PATCH'])
def api_task_priority_update(section, index):
    """Update task priority inline (for priority cycle button)."""
    data = request.get_json(force=True)
    new_priority = data.get('priority', '').strip()

    valid_priorities = ['Hi', 'Med', 'Low']
    if new_priority not in valid_priorities:
        return jsonify({'ok': False, 'error': f'priority must be one of: {", ".join(valid_priorities)}'}), 400

    lines = _read_task_lines()
    li = _find_task_line(lines, section, index)
    if li == -1:
        return jsonify({'ok': False, 'error': 'task not found'}), 404

    # Parse current task to preserve all fields
    task = _parse_task_line(lines[li], section)

    # Re-serialize with new priority
    new_line = _format_task_line(
        task['text'],
        new_priority,
        task['context'],
        task['assigned_to'],
        section,
        done=task['complete'],
        completion_date=task.get('completion_date'),
        status=task.get('status', 'New'),
        org=task.get('org', '')
    )

    lines[li] = new_line
    _write_task_file(lines)

    return jsonify({'ok': True, 'priority': new_priority})


app.register_blueprint(tasks_bp)


if __name__ == '__main__':
    app.run(port=3001, debug=True)
